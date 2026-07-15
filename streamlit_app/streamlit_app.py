# -*- coding: utf-8 -*-
"""
阿斯利康患者数据分析工具 - Streamlit 版
支持全局筛选（药房/品种）、活跃/新患统计、滚动DOT、按品种分类导出
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from io import StringIO, BytesIO
from datetime import datetime, timedelta
import re
import numpy as np


# ==================== 页面配置 ====================
st.set_page_config(
    page_title="阿斯利康患者数据分析",
    page_icon="💊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== 样式 ====================
st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1rem 0 0.5rem 0;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1.2rem 1rem;
        border-radius: 12px;
        text-align: center;
        margin-bottom: 0.5rem;
        min-height: 110px;
    }
    .metric-card .label {
        font-size: 0.85rem;
        opacity: 0.85;
        margin-bottom: 0.3rem;
    }
    .metric-card .value {
        font-size: 2rem;
        font-weight: 700;
        line-height: 1.2;
    }
    .metric-card .sub {
        font-size: 0.75rem;
        opacity: 0.7;
        margin-top: 0.2rem;
    }
    .metric-card.blue   { background: linear-gradient(135deg, #2563eb 0%, #1e40af 100%); }
    .metric-card.orange { background: linear-gradient(135deg, #ea580c 0%, #c2410c 100%); }
    .metric-card.green   { background: linear-gradient(135deg, #16a34a 0%, #15803d 100%); }
    .metric-card.purple  { background: linear-gradient(135deg, #9333ea 0%, #7e22ce 100%); }
    .metric-card.teal    { background: linear-gradient(135deg, #0891b2 0%, #0e7490 100%); }
    .metric-card.red     { background: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%); }
    .info-box {
        background: #f0f9ff;
        border: 1px solid #bae6fd;
        border-radius: 8px;
        padding: 0.8rem 1rem;
        margin: 0.5rem 0;
        font-size: 0.88rem;
        color: #075985;
    }
    .def-box {
        background: #fefce8;
        border-left: 4px solid #eab308;
        border-radius: 6px;
        padding: 0.7rem 1rem;
        margin: 0.4rem 0;
        font-size: 0.87rem;
        color: #854d0e;
    }
    .filter-section {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0;
    }
</style>
""", unsafe_allow_html=True)


# ==================== 工具函数 ====================

def parse_date(val):
    """解析各种格式的日期"""
    if pd.isna(val) or val is None or val == '':
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.Timestamp(val)
    try:
        ts = pd.to_datetime(val, errors='coerce')
        if pd.notna(ts):
            return ts
    except Exception:
        pass
    if isinstance(val, (int, float)) and 20000 < val < 100000:
        try:
            ts = pd.Timestamp('1899-12-30') + timedelta(days=float(val))
            return ts
        except Exception:
            pass
    s = str(val).strip()
    m = re.match(r'^(\d{4})[\/\-\.年](\d{1,2})', s)
    if m:
        try:
            return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=1)
        except Exception:
            pass
    return None


def get_month_key(ts):
    if ts is None:
        return None
    return f"{ts.year}-{ts.month:02d}"


def get_month_label(mk):
    if not mk or '-' not in mk:
        return str(mk)
    parts = mk.split('-')
    return f"{parts[0]}年{int(parts[1])}月"


def prev_month(mk):
    """获取上一个月的 YYYY-MM"""
    if not mk or '-' not in mk:
        return None
    parts = mk.split('-')
    y, m = int(parts[0]), int(parts[1])
    m -= 1
    if m < 1:
        m = 12
        y -= 1
    return f"{y}-{m:02d}"


def find_column(cols, keywords):
    col_lower = {c: str(c).lower().strip() for c in cols}
    for kw in keywords:
        for c, cl in col_lower.items():
            if cl == kw.lower():
                return c
    for kw in keywords:
        for c, cl in col_lower.items():
            if kw.lower() in cl:
                return c
    return None


def auto_detect_fields(columns):
    patient_keywords = ['会员号', '患者', '病人', '卡号', '会员编号', '患者id', '患者编号', 'member', 'patient', 'customer']
    date_keywords = ['销售时间', '购药时间', '购买时间', '销售日期', '购药日期', '日期', '时间', 'date', 'time']
    qty_keywords = ['销售数量', '购药数量', '数量', '购买数量', '支数', '盒数', 'qty', 'quantity', 'count']
    name_keywords = ['会员姓名', '患者姓名', '姓名', '名称', '病人姓名', 'name']
    pharmacy_keywords = ['药房名称', '药房', '药店', '门店', '门店名称', 'pharmacy', 'store', 'shop']
    product_keywords = ['商品名称', '药品名称', '品种', '产品名称', '品名', 'product', 'drug', 'medicine']

    field_map = {
        'patient': find_column(columns, patient_keywords) or (columns[0] if columns else None),
        'date': find_column(columns, date_keywords) or (columns[0] if columns else None),
        'quantity': find_column(columns, qty_keywords) or (columns[0] if columns else None),
        'name': find_column(columns, name_keywords) or None,
        'pharmacy': find_column(columns, pharmacy_keywords) or None,
        'product': find_column(columns, product_keywords) or None,
    }
    return field_map


def calculate_dropout_data(patient_data, start_month, end_month):
    """
    计算脱落率相关数据。

    定义（基于公式图片）：
    - 分析时间段 [start_month, end_month]
    - 近两月 = 结束月份及其前一个月
    - 倒推两个月前 = 分析时间段中，除「近两月」以外的其他月份
    - 脱落患者 = 在「倒推两个月前」有购药记录，但在「近两月」没有购药记录的患者
    - 脱落率 = 脱落患者数 / 倒推两个月前购药患者总人数

    返回字典包含：
    - can_compute: 是否可计算
    - dropout_rate: 百分比（如 12.34）
    - dropout_count / denominator_count
    - recent_months / prior_months
    - dropout_patients: DataFrame（序号、患者标识、患者姓名、末次购药时间、累计购药支数）
    - dropout_distribution: DataFrame（区间、人数、占比）
    """
    if not start_month or not end_month:
        return {'can_compute': False, 'reason': '缺少起止月份', 'dropout_rate': None,
                'dropout_count': 0, 'denominator_count': 0,
                'recent_months': [], 'prior_months': [],
                'dropout_patients': pd.DataFrame(), 'dropout_distribution': pd.DataFrame()}

    # 近两月
    recent_months = [end_month]
    prev = prev_month(end_month)
    if prev:
        recent_months.append(prev)

    # 分析时间段内所有月份
    period_months = []
    y, m = int(start_month.split('-')[0]), int(start_month.split('-')[1])
    ey, em = int(end_month.split('-')[0]), int(end_month.split('-')[1])
    while (y, m) <= (ey, em):
        period_months.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

    prior_months = [m for m in period_months if m not in recent_months]

    if not prior_months:
        return {'can_compute': False,
                'reason': '分析时间段不足两个月，无法计算脱落率（需至少两个完整月份）',
                'dropout_rate': None, 'dropout_count': 0, 'denominator_count': 0,
                'recent_months': recent_months, 'prior_months': prior_months,
                'dropout_patients': pd.DataFrame(), 'dropout_distribution': pd.DataFrame()}

    # 计算每个患者在分析时间段内购药的月份集合
    patient_months_in_period = {}
    for pid, p in patient_data.items():
        months = set()
        for r in p.get('records', []):
            mk = get_month_key(r.get('date'))
            if mk and start_month <= mk <= end_month:
                months.add(mk)
        patient_months_in_period[pid] = months

    # 分母：在「倒推两个月前」有购药记录的患者
    denominator = set(pid for pid, months in patient_months_in_period.items()
                      if any(m in prior_months for m in months))

    # 分子：分母中在「近两月」未购药的患者
    dropout = set(pid for pid in denominator
                  if not any(m in recent_months for m in patient_months_in_period[pid]))

    dropout_rate = (len(dropout) / len(denominator) * 100) if denominator else 0.0

    # 构建脱落患者明细
    dropout_rows = []
    for pid in dropout:
        p = patient_data[pid]
        # 只统计分析时间段内的记录
        period_records = [r for r in p.get('records', [])
                          if start_month <= get_month_key(r.get('date')) <= end_month]
        if not period_records:
            continue
        total_qty = sum(r.get('qty', 0) for r in period_records)
        last_date = max(r.get('date') for r in period_records)
        last_date_str = last_date.strftime('%Y-%m-%d') if last_date else '-'
        dropout_rows.append({
            '患者标识': pid,
            '患者姓名': p.get('name') or '-',
            '末次购药时间': last_date_str,
            '累计购药支数': round(total_qty, 2),
        })

    dropout_df = pd.DataFrame(dropout_rows)
    if not dropout_df.empty:
        dropout_df = dropout_df.sort_values('累计购药支数', ascending=False).reset_index(drop=True)
        dropout_df.insert(0, '序号', range(1, len(dropout_df) + 1))

    # 脱落患者支数分布（复用全局分布的桶）
    buckets = [('1支', 1, 1), ('2-3支', 2, 3), ('4-6支', 4, 6),
               ('7-12支', 7, 12), ('13-24支', 13, 24), ('25-36支', 25, 36), ('37支以上', 37, float('inf'))]
    bucket_data = []
    for label, lo, hi in buckets:
        if not dropout_df.empty:
            cnt = len(dropout_df[(dropout_df['累计购药支数'] >= lo) & (dropout_df['累计购药支数'] <= hi)])
        else:
            cnt = 0
        bucket_data.append({
            '购药支数区间': label,
            '患者人数': cnt,
            '占比(%)': round(cnt / len(dropout) * 100, 2) if dropout else 0,
        })

    # ==================== 每月滚动脱落率趋势（新增脱落 / 基准月 T-2） ====================
    # 对回顾月份 M：基准月 T-2 = M 的前两个月
    #   分母   = 仅在基准月 T-2 有购药的患者（当月活跃基准人群，单月）
    #   脱落   = 分母中在 T-1、T（即 M-1、M）都未购药的患者（即当月新增脱落）
    #   脱落率 = 脱落 / 分母 × 100%
    # 说明：分母只取基准月 T-2 单月，不是"T-2 及更早"的累计；由于分母随 M 前移，
    #       同一患者只在「首次满足脱落条件」的那个月（即其最后一次购药月 +2）被计数一次，天然即为"新增"语义。
    trend_rows = []
    for i, cur_m in enumerate(period_months):
        baseline = prev_month(prev_month(cur_m))  # T-2
        if baseline is None or baseline not in period_months:
            continue  # 基准月不在分析时间段内（最早两个月无法回顾），跳过
        # 分母：仅在基准月 T-2 有购药的患者
        cur_denom = set(pid for pid, months in patient_months_in_period.items()
                        if baseline in months)
        if not cur_denom:
            continue
        # 脱落：分母中在 M-1、M 都未购药的患者
        recent = {cur_m}
        pm = prev_month(cur_m)
        if pm:
            recent.add(pm)
        cur_drop = set(pid for pid in cur_denom
                       if not any(m in recent for m in patient_months_in_period[pid]))
        cur_rate = len(cur_drop) / len(cur_denom) * 100
        trend_rows.append({
            '回顾月份': get_month_label(cur_m),
            '_month_key': cur_m,
            '脱落人数': len(cur_drop),
            '分母人数': len(cur_denom),
            '脱落率(%)': round(cur_rate, 2),
        })
    monthly_trend_df = pd.DataFrame(trend_rows)

    return {
        'can_compute': True,
        'reason': '',
        'dropout_rate': round(dropout_rate, 2),
        'dropout_count': len(dropout),
        'denominator_count': len(denominator),
        'recent_months': recent_months,
        'prior_months': prior_months,
        'dropout_patients': dropout_df,
        'dropout_distribution': pd.DataFrame(bucket_data),
        'monthly_trend': monthly_trend_df,
    }


# ==================== 初始化 Session State ====================
for key in [
    'raw_df', 'columns', 'field_map', 'available_months',
    'dot_result', 'analysis_data', 'annual_comparison'
]:
    if key not in st.session_state:
        st.session_state[key] = None


# ==================== 标题 ====================
st.markdown("""
<div class="main-header">
    <h1>💊 阿斯利康患者数据分析</h1>
    <p style="color: #64748b; font-size: 1.05rem;">DOT 用药时长分析 · 活跃/新患统计 · 多维度数据可视化 · 支持药房/品种筛选</p>
</div>
""", unsafe_allow_html=True)


# ==================== 侧边栏 ====================
with st.sidebar:
    st.header("📊 数据导入")
    uploaded_file = st.file_uploader(
        "上传销售数据底表",
        type=['xlsx', 'xls', 'csv'],
        help="支持 Excel (.xlsx, .xls) 和 CSV 格式"
    )

    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            if df.empty:
                st.error("文件中没有数据")
            else:
                st.session_state.raw_df = df
                st.session_state.columns = list(df.columns)
                detected = auto_detect_fields(list(df.columns))
                st.session_state.field_map = detected

                date_col = detected.get('date')
                if date_col:
                    months = set()
                    for val in df[date_col]:
                        ts = parse_date(val)
                        if ts:
                            months.add(get_month_key(ts))
                    st.session_state.available_months = sorted(months)

                st.success(f"✅ 数据加载成功！共 {len(df):,} 条记录，{len(df.columns)} 个字段")

                with st.expander("📁 文件信息", expanded=False):
                    st.write(f"**文件名：** {uploaded_file.name}")
                    st.write(f"**记录数：** {len(df):,}")
                    st.write(f"**字段列表：** {', '.join(str(c) for c in df.columns)}")

        except Exception as e:
            st.error(f"文件解析失败: {str(e)}")

    st.markdown("---")
    st.markdown("""
    ### 📖 指标说明

    **活跃患者数**
    > 在所选时间段内**至少有 1 次购药记录**的去重患者人数。

    **新患患者数（月度）**
    > 在该月份购药的患者中，**首次出现在全量数据中的**去重患者人数。
    > 即：该患者的第一条购药记录就在这个月。

    **DOT 值**
    ```
    DOT = 累计购药总支数 / 去重患者数
    ```

    **滚动一年算法（月度 DOT）**
    > 以某月为基准，向前滚动 12 个月作为窗口期：
    > 例：2026年1月的 DOT = 2025年2月~2026年1月期间的去重患者总支数 / 去重患者数
    """)


# ==================== 主内容区 ====================
if st.session_state.raw_df is None:
    st.markdown("""
    <div style="text-align: center; padding: 3rem 1rem;">
        <div style="font-size: 4rem; margin-bottom: 1rem;">📋</div>
        <h3 style="color: #475569;">请在左侧上传销售数据底表开始分析</h3>
        <p style="color: #94a3b8; max-width: 500px; margin: 0 auto;">
            支持 Excel (.xlsx, .xls) 和 CSV 格式文件。<br/>
            所有数据在本地处理，不上传至服务器。
        </p>
    </div>
    """, unsafe_allow_html=True)
else:
    df = st.session_state.raw_df
    columns = st.session_state.columns
    field_map = st.session_state.field_map

    # ==================== 字段映射 ====================
    st.header("⚙️ 字段配置")
    st.caption("系统已自动识别字段，如有误可手动调整")

    col1, col2, col3 = st.columns(3)
    with col1:
        patient_col = st.selectbox("患者标识 *", options=columns,
            index=columns.index(field_map['patient']) if field_map.get('patient') in columns else 0, key='patient_field')
    with col2:
        date_col = st.selectbox("购药时间 *", options=columns,
            index=columns.index(field_map['date']) if field_map.get('date') in columns else 0, key='date_field')
    with col3:
        qty_col = st.selectbox("购药数量 *", options=columns,
            index=columns.index(field_map['quantity']) if field_map.get('quantity') in columns else 0, key='qty_field')

    col4, col5, col6 = st.columns(3)
    with col4:
        no = ['（不使用）'] + columns
        ni = 0
        if field_map.get('name') and field_map['name'] in columns:
            ni = no.index(field_map['name'])
        name_sel = st.selectbox("患者姓名（可选）", options=no, index=ni, key='name_field')
        name_col = None if name_sel == '（不使用）' else name_sel
    with col5:
        po = ['（不使用）'] + columns
        pi = 0
        if field_map.get('pharmacy') and field_map['pharmacy'] in columns:
            pi = po.index(field_map['pharmacy'])
        pharma_sel = st.selectbox("药房名称（可选）", options=po, index=pi, key='pharmacy_field')
        pharmacy_col = None if pharma_sel == '（不使用）' else pharma_sel
    with col6:
        pro_o = ['（不使用）'] + columns
        pro_i = 0
        if field_map.get('product') and field_map['product'] in columns:
            pro_i = pro_o.index(field_map['product'])
        prod_sel = st.selectbox("品种/商品名称（可选）", options=pro_o, index=pro_i, key='product_field')
        product_col = None if prod_sel == '（不使用）' else prod_sel

    # 更新可用月份
    if date_col:
        months = set()
        for val in df[date_col]:
            ts = parse_date(val)
            if ts:
                months.add(get_month_key(ts))
        st.session_state.available_months = sorted(months)

    # ==================== 全局筛选条件 ====================
    available_months = st.session_state.available_months

    if not available_months:
        st.warning("⚠️ 未检测到有效的日期数据，请检查「购药时间」字段是否正确")
    else:
        month_labels = [get_month_label(m) for m in available_months]

        st.markdown("---")
        st.header("🔍 全局筛选条件")
        st.caption("以下筛选条件将影响所有图表和数据展示")

        st.markdown('<div class="filter-section">', unsafe_allow_html=True)

        fc1, fc2 = st.columns(2)
        with fc1:
            si = st.selectbox("起始月份", options=range(len(available_months)),
                format_func=lambda i: month_labels[i], index=0, key='start_month')
        with fc2:
            ei_val = len(available_months) - 1
            ei = st.selectbox("结束月份", options=range(len(available_months)),
                format_func=lambda i: month_labels[i], index=ei_val, key='end_month')

        if si <= ei:
            start_month = available_months[si]
            end_month = available_months[ei]
        else:
            start_month = end_month = None
            st.error("起始月份不能晚于结束月份")

        # 药房筛选
        selected_pharmacies = ['全部']
        if pharmacy_col:
            all_pharmacies = sorted([str(v).strip() for v in df[pharmacy_col].dropna().unique()])
            ph_mode = st.radio("药房筛选模式", ["全部药房", "选择单个药房", "多选药房"], horizontal=True, key='ph_mode')
            if ph_mode == "选择单个药房":
                sel_ph = st.selectbox("选择药房", options=all_pharmacies, key='sel_pharmacy_single')
                selected_pharmacies = [sel_ph]
            elif ph_mode == "多选药房":
                sel_multi = st.multiselect("选择药房（可多选）", options=all_pharmacies, default=all_pharmacies[:3], key='sel_pharmacy_multi')
                selected_pharmacies = sel_multi if sel_multi else ['全部']
            else:
                selected_pharmacies = ['全部']

        # 品种筛选
        selected_products = ['全部']
        if product_col:
            all_products = sorted([str(v).strip() for v in df[product_col].dropna().unique()])
            sel_prod = st.multiselect("品种筛选（可多选）", options=all_products, default=all_products, key='sel_products')
            selected_products = sel_prod if sel_prod else ['全部']

        st.markdown('</div>', unsafe_allow_html=True)

        # 筛选信息摘要
        ph_text = ', '.join(selected_pharmacies[:5]) + (f' ... (+{len(selected_pharmacies)-5})' if len(selected_pharmacies) > 5 else '')
        pr_text = ', '.join(selected_products[:3]) + (f' ... (+{len(selected_products)-3})' if len(selected_products) > 3 else '')

        if start_month and end_month:
            info_parts = [f"时间：{get_month_label(start_month)} ~ {get_month_label(end_month)}"]
            if pharmacy_col and selected_pharmacies != ['全部']:
                info_parts.append(f"药房：{ph_text}")
            if product_col and selected_products != ['全部']:
                info_parts.append(f"品种：{pr_text}")
            st.info("**当前筛选：** " + " | ".join(info_parts))

            # ==================== 分析按钮 ====================
            if st.button("🚀 开始分析", type="primary", use_container_width=True):
                with st.spinner("正在计算指标..."):
                    # ---- 构建筛选后的 DataFrame ----
                    filtered_df = df.copy()

                    if pharmacy_col and selected_pharmacies != ['全部']:
                        filtered_df = filtered_df[filtered_df[pharmacy_col].astype(str).str.strip().isin(selected_pharmacies)]

                    if product_col and selected_products != ['全部']:
                        filtered_df = filtered_df[filtered_df[product_col].astype(str).str.strip().isin(selected_products)]

                    if filtered_df.empty:
                        st.error("筛选后没有数据，请调整筛选条件")
                    else:
                        # ---- Step 1: 找出时间段内有购药记录的患者 ----
                        patients_in_period = set()
                        for _, row in filtered_df.iterrows():
                            ts = parse_date(row[date_col])
                            if ts is None:
                                continue
                            mk = get_month_key(ts)
                            if start_month <= mk <= end_month:
                                pid = str(row[patient_col]).strip()
                                if pid and pid != 'nan':
                                    patients_in_period.add(pid)

                        unique_patient_count = len(patients_in_period)

                        if unique_patient_count == 0:
                            st.error("所选时间段内没有购药记录")
                        else:
                            # ---- Step 2: 统计这些患者从首次到最新的全部购药 ----
                            patient_data = {}

                            for _, row in filtered_df.iterrows():
                                pid = str(row[patient_col]).strip()
                                if pid not in patients_in_period or pid == 'nan':
                                    continue
                                ts = parse_date(row[date_col])
                                if ts is None:
                                    continue
                                try:
                                    qty = float(row[qty_col])
                                except (ValueError, TypeError):
                                    qty = 0

                                pharm_name = ''
                                if pharmacy_col:
                                    pv = row.get(pharmacy_col, '')
                                    if pd.notna(pv):
                                        pharm_name = str(pv).strip()

                                prod_name = ''
                                if product_col:
                                    pv2 = row.get(product_col, '')
                                    if pd.notna(pv2):
                                        prod_name = str(pv2).strip()

                                if pid not in patient_data:
                                    patient_data[pid] = {
                                        'id': pid, 'name': '', 'records': [],
                                        'total_qty': 0, 'total_visits': 0,
                                        'first_date': ts, 'last_date': ts,
                                        'pharmacies': set(),
                                        'pharmacy_qty': {}, 'pharmacy_visits': {},
                                        'products': set(), 'product_qty': {},
                                    }

                                p = patient_data[pid]
                                p['records'].append({'date': ts, 'qty': qty, 'pharmacy': pharm_name, 'product': prod_name})
                                p['total_qty'] += qty
                                p['total_visits'] += 1
                                if ts < p['first_date']: p['first_date'] = ts
                                if ts > p['last_date']: p['last_date'] = ts

                                if not p['name'] and name_col:
                                    nv = row.get(name_col, '')
                                    if pd.notna(nv): p['name'] = str(nv).strip()

                                if pharm_name:
                                    p['pharmacies'].add(pharm_name)
                                    p['pharmacy_qty'][pharm_name] = p['pharmacy_qty'].get(pharm_name, 0) + qty
                                    p['pharmacy_visits'][pharm_name] = p['pharmacy_visits'].get(pharm_name, 0) + 1
                                if prod_name:
                                    p['products'].add(prod_name)
                                    p['product_qty'][prod_name] = p['product_qty'].get(prod_name, 0) + qty

                            total_quantity = sum(p['total_qty'] for p in patient_data.values())
                            dot_value = total_quantity / unique_patient_count if unique_patient_count > 0 else 0

                            # ---- 患者明细 DataFrame ----
                            detail_rows = []
                            for p in patient_data.values():
                                detail_rows.append({
                                    '患者标识': p['id'], '患者姓名': p['name'] or '-',
                                    '药房数': len(p['pharmacies']),
                                    '主要药房': ', '.join(sorted(p['pharmacies'])) if p['pharmacies'] else '-',
                                    '首次购药时间': p['first_date'].strftime('%Y-%m-%d') if p['first_date'] else '-',
                                    '末次购药时间': p['last_date'].strftime('%Y-%m-%d') if p['last_date'] else '-',
                                    '购药总次数': p['total_visits'],
                                    '购药总支数': round(p['total_qty'], 2),
                                })

                            details_df = pd.DataFrame(detail_rows)
                            details_df = details_df.sort_values('购药总支数', ascending=False).reset_index(drop=True)
                            details_df.insert(0, '序号', range(1, len(details_df) + 1))

                            # ---- 月度统计数据（用于所有图表）----
                            monthly_stats = {}
                            for _, row in filtered_df.iterrows():
                                pid = str(row[patient_col]).strip()
                                if pid not in patients_in_period or pid == 'nan':
                                    continue
                                ts = parse_date(row[date_col])
                                if ts is None:
                                    continue
                                mk = get_month_key(ts)
                                if mk not in monthly_stats:
                                    monthly_stats[mk] = {'qty': 0, 'patients': set()}
                                try:
                                    qv = float(row[qty_col])
                                except (ValueError, TypeError):
                                    qv = 0
                                monthly_stats[mk]['qty'] += qv
                                monthly_stats[mk]['patients'].add(pid)

                            # ---- 新患统计（每月首次出现的患者）----
                            # 需要基于全量数据计算每个患者的首次购药月份
                            patient_first_month = {}
                            for _, row in filtered_df.iterrows():
                                pid = str(row[patient_col]).strip()
                                if pid == 'nan' or not pid:
                                    continue
                                if pid not in patients_in_period:
                                    continue
                                ts = parse_date(row[date_col])
                                if ts is None:
                                    continue
                                mk = get_month_key(ts)
                                if pid not in patient_first_month or mk < patient_first_month[pid]:
                                    patient_first_month[pid] = mk

                            # 新患：按月统计 first_month == 该月的患者数
                            new_patient_by_month = {}
                            for pid, fmk in patient_first_month.items():
                                if fmk not in new_patient_by_month:
                                    new_patient_by_month[fmk] = set()
                                new_patient_by_month[fmk].add(pid)

                            # ---- 药房汇总 ----
                            pharmacy_summary_rows = []
                            if pharmacy_col:
                                pharm_stats = {}
                                for _, row in filtered_df.iterrows():
                                    pid = str(row[patient_col]).strip()
                                    if pid not in patients_in_period or pid == 'nan':
                                        continue
                                    ts = parse_date(row[date_col])
                                    if ts is None:
                                        continue
                                    pv = row.get(pharmacy_col, '')
                                    if pd.isna(pv): continue
                                    pn = str(pv).strip()
                                    if not pn: continue
                                    try: q = float(row[qty_col])
                                    except (ValueError, TypeError): q = 0
                                    if pn not in pharm_stats:
                                        pharm_stats[pn] = {'patients': set(), 'qty': 0, 'visits': 0}
                                    pharm_stats[pn]['patients'].add(pid)
                                    pharm_stats[pn]['qty'] += q
                                    pharm_stats[pn]['visits'] += 1

                                for pn, ps in sorted(pharm_stats.items(), key=lambda x: x[1]['qty'], reverse=True):
                                    pc = len(ps['patients'])
                                    pharmacy_summary_rows.append({
                                        '药房名称': pn, '去重患者数': pc,
                                        '购药总次数': ps['visits'],
                                        '累计购药总支数': round(ps['qty'], 2),
                                        'DOT值': round(ps['qty'] / pc, 2) if pc > 0 else 0,
                                    })

                            pharmacy_summary_df = pd.DataFrame(pharmacy_summary_rows)
                            if not pharmacy_summary_df.empty:
                                pharmacy_summary_df.insert(0, '序号', range(1, len(pharmacy_summary_df) + 1))

                            # ---- 品种汇总 ----
                            product_summary_rows = []
                            if product_col:
                                prod_stats = {}
                                for _, row in filtered_df.iterrows():
                                    pid = str(row[patient_col]).strip()
                                    if pid not in patients_in_period or pid == 'nan':
                                        continue
                                    ts = parse_date(row[date_col])
                                    if ts is None: continue
                                    pv = row.get(product_col, '')
                                    if pd.isna(pv): continue
                                    pn = str(pv).strip()
                                    if not pn: continue
                                    try: q = float(row[qty_col])
                                    except (ValueError, TypeError): q = 0
                                    if pn not in prod_stats:
                                        prod_stats[pn] = {'patients': set(), 'qty': 0, 'visits': 0}
                                    prod_stats[pn]['patients'].add(pid)
                                    prod_stats[pn]['qty'] += q
                                    prod_stats[pn]['visits'] += 1

                                for pn, ps in sorted(prod_stats.items(), key=lambda x: x[1]['qty'], reverse=True):
                                    pc = len(ps['patients'])
                                    product_summary_rows.append({
                                        '品种': pn, '去重患者数': pc,
                                        '购药总次数': ps['visits'],
                                        '累计购药总支数': round(ps['qty'], 2),
                                        'DOT值': round(ps['qty'] / pc, 2) if pc > 0 else 0,
                                    })

                            product_summary_df = pd.DataFrame(product_summary_rows)
                            if not product_summary_df.empty:
                                product_summary_df.insert(0, '序号', range(1, len(product_summary_df) + 1))

                            # ---- 按 品种×药房 汇总 ----
                            cross_summary_rows = []
                            if product_col and pharmacy_col:
                                cross_stats = {}
                                for _, row in filtered_df.iterrows():
                                    pid = str(row[patient_col]).strip()
                                    if pid not in patients_in_period or pid == 'nan':
                                        continue
                                    ts = parse_date(row[date_col])
                                    if ts is None: continue
                                    pv1 = row.get(product_col, '')
                                    pv2 = row.get(pharmacy_col, '')
                                    if pd.isna(pv1) or pd.isna(pv2): continue
                                    pn1 = str(pv1).strip()
                                    pn2 = str(pv2).strip()
                                    if not pn1 or not pn2: continue
                                    try: q = float(row[qty_col])
                                    except (ValueError, TypeError): q = 0
                                    key = (pn1, pn2)
                                    if key not in cross_stats:
                                        cross_stats[key] = {'patients': set(), 'qty': 0, 'visits': 0}
                                    cross_stats[key]['patients'].add(pid)
                                    cross_stats[key]['qty'] += q
                                    cross_stats[key]['visits'] += 1

                                for (pn1, pn2), cs in sorted(cross_stats.items(), key=lambda x: (x[0][0], -x[1]['qty'])):
                                    pc = len(cs['patients'])
                                    cross_summary_rows.append({
                                        '品种': pn1, '药房名称': pn2,
                                        '去重患者数': pc, '购药总次数': cs['visits'],
                                        '累计购药总支数': round(cs['qty'], 2),
                                        'DOT值': round(cs['qty'] / pc, 2) if pc > 0 else 0,
                                    })

                            cross_summary_df = pd.DataFrame(cross_summary_rows)
                            if not cross_summary_df.empty:
                                cross_summary_df.insert(0, '序号', range(1, len(cross_summary_df) + 1))

                            # ---- 按 品种×药房 的患者明细 ----
                            cross_detail_rows = []
                            if product_col and pharmacy_col:
                                for p in patient_data.values():
                                    for prod in sorted(p['products']):
                                        for pharm in sorted(p['pharmacies']):
                                            # 计算该患者在此品种+药房组合下的数据
                                            pqty = 0
                                            pvisits = 0
                                            pf = pl = None
                                            for r in p['records']:
                                                if r.get('product') == prod and r.get('pharmacy') == pharm:
                                                    pqty += r['qty']
                                                    pvisits += 1
                                                    if pf is None or r['date'] < pf: pf = r['date']
                                                    if pl is None or r['date'] > pl: pl = r['date']
                                            if pqty > 0:
                                                cross_detail_rows.append({
                                                    '品种': prod, '药房名称': pharm,
                                                    '患者标识': p['id'],
                                                    '患者姓名': p['name'] or '-',
                                                    '首次购药时间': pf.strftime('%Y-%m-%d') if pf else '-',
                                                    '末次购药时间': pl.strftime('%Y-%m-%d') if pl else '-',
                                                    '购药次数': pvisits,
                                                    '累计购药支数': round(pqty, 2),
                                                })

                            cross_detail_df = pd.DataFrame(cross_detail_rows)
                            if not cross_detail_df.empty:
                                cross_detail_df = cross_detail_df.sort_values(
                                    ['品种', '药房名称', '累计购药支数'], ascending=[True, True, False]
                                ).reset_index(drop=True)
                                cross_detail_df.insert(0, '序号', range(1, len(cross_detail_df) + 1))

                            st.session_state.dot_result = {
                            'start_month': start_month, 'end_month': end_month,
                            'unique_patient_count': unique_patient_count,
                            'total_quantity': round(total_quantity, 2),
                            'dot_value': round(dot_value, 2),
                            'patient_ids': patients_in_period,
                            'selected_pharmacies': selected_pharmacies,
                            'selected_products': selected_products,
                            'pharmacy_col': pharmacy_col, 'product_col': product_col,
                        }

                        # ---- 年度DOT对比计算 ----
                        # 同期年度对比：去年同期
                        def parse_month_key(mk):
                            if not mk or '-' not in mk:
                                return None
                            parts = mk.split('-')
                            return (int(parts[0]), int(parts[1]))

                        def month_key_to_str(mk):
                            return mk

                        def get_year_month_range(start_mk, end_mk):
                            """获取从start_mk到end_mk的所有月份"""
                            result = []
                            y, m = parse_month_key(start_mk)
                            ey, em = parse_month_key(end_mk)
                            while (y, m) <= (ey, em):
                                result.append(f"{y}-{m:02d}")
                                m += 1
                                if m > 12:
                                    m = 1
                                    y += 1
                            return result

                        # 同期年度
                        sm_y, sm_m = parse_month_key(start_month)
                        em_y, em_m = parse_month_key(end_month)
                        same_period_last_year_start = f"{sm_y-1}-{sm_m:02d}"
                        same_period_last_year_end = f"{em_y-1}-{em_m:02d}"

                        # 滚动年度：基于结束月份往前12个月
                        def get_rolling_12_months(end_mk):
                            """获取以end_mk为终点的前12个月"""
                            y, m = parse_month_key(end_mk)
                            months = []
                            for _ in range(12):
                                months.append(f"{y}-{m:02d}")
                                m -= 1
                                if m < 1:
                                    m = 12
                                    y -= 1
                            return list(reversed(months))

                        rolling_current = get_rolling_12_months(end_month)
                        rolling_current_start = rolling_current[0]
                        rolling_current_end = rolling_current[-1]

                        # 上一个滚动年度
                        ry, rm = parse_month_key(rolling_current_start)
                        rolling_prev_start = f"{ry-1}-{rm:02d}"
                        ry2, rm2 = parse_month_key(rolling_current_end)
                        rolling_prev_end = f"{ry2-1}-{rm2:02d}"

                        # 计算同期年度数据
                        def calc_period_dot(filtered_df, period_start, period_end, patient_col, date_col, qty_col):
                            """计算某个时间段的DOT数据"""
                            period_patients = set()
                            for _, row in filtered_df.iterrows():
                                ts = parse_date(row[date_col])
                                if ts is None:
                                    continue
                                mk = get_month_key(ts)
                                if period_start <= mk <= period_end:
                                    pid = str(row[patient_col]).strip()
                                    if pid and pid != 'nan':
                                        period_patients.add(pid)

                            if not period_patients:
                                return {'patients': 0, 'total_qty': 0, 'dot': 0, 'patient_ids': set()}

                            # 统计这些患者的全部购药
                            total_qty = 0
                            for _, row in filtered_df.iterrows():
                                pid = str(row[patient_col]).strip()
                                if pid not in period_patients or pid == 'nan':
                                    continue
                                try:
                                    qty = float(row[qty_col])
                                except (ValueError, TypeError):
                                    qty = 0
                                total_qty += qty

                            patient_count = len(period_patients)
                            return {
                                'patients': patient_count,
                                'total_qty': round(total_qty, 2),
                                'dot': round(total_qty / patient_count, 2) if patient_count > 0 else 0,
                                'patient_ids': period_patients
                            }

                        # 同期年度数据
                        same_period_current = calc_period_dot(filtered_df, start_month, end_month, patient_col, date_col, qty_col)
                        same_period_last = calc_period_dot(filtered_df, same_period_last_year_start, same_period_last_year_end, patient_col, date_col, qty_col)

                        # 滚动年度数据
                        rolling_current_data = calc_period_dot(filtered_df, rolling_current_start, rolling_current_end, patient_col, date_col, qty_col)
                        rolling_prev_data = calc_period_dot(filtered_df, rolling_prev_start, rolling_prev_end, patient_col, date_col, qty_col)

                        st.session_state.annual_comparison = {
                            'mode': 'same_period',  # 默认同期对比
                            'same_period': {
                                'current': same_period_current,
                                'last': same_period_last,
                                'current_label': f"{get_month_label(start_month)}~{get_month_label(end_month)}",
                                'last_label': f"{get_month_label(same_period_last_year_start)}~{get_month_label(same_period_last_year_end)}",
                            },
                            'rolling': {
                                'current': rolling_current_data,
                                'prev': rolling_prev_data,
                                'current_label': f"{get_month_label(rolling_current_start)}~{get_month_label(rolling_current_end)}",
                                'prev_label': f"{get_month_label(rolling_prev_start)}~{get_month_label(rolling_prev_end)}",
                            }
                        }

                        # 预计算图表用数据（供导出和展示复用）
                        all_months_sorted = sorted(monthly_stats.keys())
                        ml_all = [get_month_label(m) for m in all_months_sorted]
                        qty_all = [round(monthly_stats[m]['qty'], 2) for m in all_months_sorted]
                        active_pats_all = [len(monthly_stats[m]['patients']) for m in all_months_sorted]
                        new_pats_all = [len(new_patient_by_month.get(m, set())) for m in all_months_sorted]
                        bar_colors = []
                        for m in all_months_sorted:
                            if start_month <= m <= end_month:
                                bar_colors.append('rgba(37, 99, 235, 0.78)')
                            else:
                                bar_colors.append('rgba(148, 163, 184, 0.35)')

                        # ---- 按药房维度：月度统计 + 新患统计 ----
                        pharmacy_monthly_data = {}  # {pharmacy_name: {month_key: {qty, patients_set}}}
                        pharmacy_new_patient_month = {}  # {pharmacy_name: {month_key: first_month_in_this_pharm}}
                        if pharmacy_col:
                            # Step1: 建立每个药房下，每个患者的最早购药月份
                            pharm_patient_first_month = {}  # {pharmacy: {pid: first_month}}
                            for _, row in filtered_df.iterrows():
                                pid = str(row[patient_col]).strip()
                                if pid not in patients_in_period or pid == 'nan':
                                    continue
                                ts = parse_date(row[date_col])
                                if ts is None:
                                    continue
                                pv = row.get(pharmacy_col, '')
                                if pd.isna(pv):
                                    continue
                                pname = str(pv).strip()
                                if not pname:
                                    continue
                                mk = get_month_key(ts)
                                try:
                                    qv = float(row[qty_col])
                                except (ValueError, TypeError):
                                    qv = 0
                                # 月度统计
                                if pname not in pharmacy_monthly_data:
                                    pharmacy_monthly_data[pname] = {}
                                if mk not in pharmacy_monthly_data[pname]:
                                    pharmacy_monthly_data[pname][mk] = {'qty': 0, 'patients': set()}
                                pharmacy_monthly_data[pname][mk]['qty'] += qv
                                pharmacy_monthly_data[pname][mk]['patients'].add(pid)
                                # 新患统计：该患者在该药房的最早月份
                                if pname not in pharm_patient_first_month:
                                    pharm_patient_first_month[pname] = {}
                                if pid not in pharm_patient_first_month[pname] or mk < pharm_patient_first_month[pname][pid]:
                                    pharm_patient_first_month[pname][pid] = mk
                            # 汇总成 {pharmacy: {month_key: patients_set}}
                            for pname, pid_map in pharm_patient_first_month.items():
                                pharmacy_new_patient_month[pname] = {}
                                for pid, fmk in pid_map.items():
                                    if fmk not in pharmacy_new_patient_month[pname]:
                                        pharmacy_new_patient_month[pname][fmk] = set()
                                    pharmacy_new_patient_month[pname][fmk].add(pid)

                        st.session_state.analysis_data = {
                            'details_df': details_df,
                            'pharmacy_summary_df': pharmacy_summary_df,
                            'product_summary_df': product_summary_df,
                            'cross_summary_df': cross_summary_df,
                            'cross_detail_df': cross_detail_df,
                            'monthly_stats': monthly_stats,
                            'new_patient_by_month': new_patient_by_month,
                            'patient_first_month': patient_first_month,
                            'patient_data': patient_data,
                            'all_months_sorted': all_months_sorted,
                            'ml_all': ml_all,
                            'qty_all': qty_all,
                            'active_pats_all': active_pats_all,
                            'new_pats_all': new_pats_all,
                            'bar_colors': bar_colors,
                            'pharmacy_monthly_data': pharmacy_monthly_data,
                            'pharmacy_new_patient_month': pharmacy_new_patient_month,
                            'pharmacy_col_name': pharmacy_col,
                            'start_month': start_month,
                            'end_month': end_month,
                            'filtered_df': filtered_df,
                            'date_col': date_col,
                            'patient_col': patient_col,
                            'qty_col': qty_col,
                        }

                        # ---- 脱落率分析 ----
                        dropout_data = calculate_dropout_data(
                            patient_data, start_month, end_month
                        )
                        st.session_state.analysis_data['dropout_data'] = dropout_data

    # ==================== 结果展示 ====================
    if st.session_state.dot_result is not None:
        result = st.session_state.dot_result
        ad = st.session_state.analysis_data
        # 提前提示：未识别到日期列时，导出 Excel 必然失败（避免用户等到导出才崩）
        if not ad.get('date_col'):
            st.warning("⚠️ 本次分析未识别到『购药时间』日期列：图表仍可正常查看，但【导出 Excel】会失败。"
                       "请返回上方『字段配置』重新选择『购药时间 *』，再点击「运行分析」后导出。")
        details_df = ad['details_df']
        monthly_stats = ad['monthly_stats']
        new_patient_by_month = ad['new_patient_by_month']
        patient_first_month = ad['patient_first_month']

        st.markdown("---")
        st.header("📊 分析结果")

        # ---- 定义说明框 ----
        st.markdown("""
        <div class="def-box">
        <strong>📖 活跃患者数定义：</strong>在所选时间段内至少有 1 次购药记录的去重患者人数。<br/>
        统计范围覆盖这些患者从<strong>首次购药</strong>到<strong>最新销售数据</strong>为止的全部购药记录。
        </div>
        """, unsafe_allow_html=True)

        # 核心指标卡片
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f"""
            <div class="metric-card purple">
                <div class="label">筛选时间段</div>
                <div class="value" style="font-size: 1rem;">{get_month_label(result['start_month'])}<br/>~ {get_month_label(result['end_month'])}</div>
            </div>
            """, unsafe_allow_html=True)
        with c2:
            st.markdown(f"""
            <div class="metric-card blue">
                <div class="label">活跃患者数</div>
                <div class="value">{result['unique_patient_count']:,}</div>
                <div class="sub">时间段内有购药记录</div>
            </div>
            """, unsafe_allow_html=True)
        with c3:
            st.markdown(f"""
            <div class="metric-card orange">
                <div class="label">累计购药总支数</div>
                <div class="value">{result['total_quantity']:,.1f}</div>
                <div class="sub">全时段累计</div>
            </div>
            """, unsafe_allow_html=True)
        with c4:
            st.markdown(f"""
            <div class="metric-card green">
                <div class="label">DOT 值</div>
                <div class="value">{result['dot_value']:.2f}</div>
                <div class="sub">= 总支数 / 活跃患者数</div>
            </div>
            """, unsafe_allow_html=True)
        with c5:
            # 计算时间段内的总新患数
            total_new_patients = 0
            for mk in sorted(monthly_stats.keys()):
                if result['start_month'] <= mk <= result['end_month']:
                    total_new_patients += len(new_patient_by_month.get(mk, set()))
            st.markdown(f"""
            <div class="metric-card teal">
                <div class="label">新患患者数</div>
                <div class="value">{total_new_patients:,}</div>
                <div class="sub">{get_month_label(result['start_month'])}~{get_month_label(result['end_month'])}期间首购</div>
            </div>
            """, unsafe_allow_html=True)

        # ==================== 年度DOT对比 ====================
        st.markdown("---")
        st.header("📊 年度 DOT 对比分析")

        if st.session_state.annual_comparison is not None:
            ac = st.session_state.annual_comparison

            # 切换选项
            comp_mode = st.radio(
                "对比模式",
                ["同期年度DOT对比", "滚动年度DOT对比"],
                horizontal=True,
                key='annual_comp_mode'
            )

            if comp_mode == "同期年度DOT对比":
                st.markdown("""
                <div class="def-box" style="font-size:0.82rem;">
                <strong>同期年度DOT对比：</strong>将所选时间段与去年同期的DOT数据进行对比。<br/>
                例：选择 <strong>2026年1月~6月</strong>，则对比 <strong>2025年1月~6月</strong> 的数据。
                </div>
                """, unsafe_allow_html=True)

                col_a, col_b = st.columns(2)

                with col_a:
                    st.subheader(f"📅 当前年度：{ac['same_period']['current_label']}")
                    curr = ac['same_period']['current']
                    st.markdown(f"""
                    <div style="background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <p style="margin: 0.3rem 0;"><strong>去重患者数：</strong>{curr['patients']:,} 人</p>
                        <p style="margin: 0.3rem 0;"><strong>累计购药总支数：</strong>{curr['total_qty']:,.1f} 支</p>
                        <p style="margin: 0.3rem 0;"><strong>DOT 值：</strong><span style="font-size: 1.3rem; font-weight: 700; color: #2563eb;">{curr['dot']:.2f}</span></p>
                    </div>
                    """, unsafe_allow_html=True)

                with col_b:
                    st.subheader(f"📅 去年同期：{ac['same_period']['last_label']}")
                    last = ac['same_period']['last']
                    if last['patients'] > 0:
                        dot_change = ((curr['dot'] - last['dot']) / last['dot'] * 100) if last['dot'] > 0 else 0
                        patient_change = ((curr['patients'] - last['patients']) / last['patients'] * 100) if last['patients'] > 0 else 0
                        qty_change = ((curr['total_qty'] - last['total_qty']) / last['total_qty'] * 100) if last['total_qty'] > 0 else 0

                        st.markdown(f"""
                        <div style="background: #fefce8; border: 1px solid #fde68a; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                            <p style="margin: 0.3rem 0;"><strong>去重患者数：</strong>{last['patients']:,} 人
                            <span style="color: {'#16a34a' if patient_change >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if patient_change >= 0 else '↓'}{abs(patient_change):.1f}%)</span></p>
                            <p style="margin: 0.3rem 0;"><strong>累计购药总支数：</strong>{last['total_qty']:,.1f} 支
                            <span style="color: {'#16a34a' if qty_change >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if qty_change >= 0 else '↓'}{abs(qty_change):.1f}%)</span></p>
                            <p style="margin: 0.3rem 0;"><strong>DOT 值：</strong><span style="font-size: 1.3rem; font-weight: 700; color: #ea580c;">{last['dot']:.2f}</span>
                            <span style="color: {'#16a34a' if dot_change >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if dot_change >= 0 else '↓'}{abs(dot_change):.1f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("去年同期无数据")

                # 对比图表
                st.subheader("同期年度DOT对比图表")

                compare_labels = [ac['same_period']['current_label'], ac['same_period']['last_label']]
                compare_patients = [curr['patients'], last['patients']]
                compare_qty = [curr['total_qty'], last['total_qty']]
                compare_dot = [curr['dot'], last['dot']]

                fig_annual = make_subplots(specs=[[{"secondary_y": True}]])

                fig_annual.add_trace(go.Bar(
                    x=compare_labels, y=compare_qty, name='累计购药总支数',
                    marker_color=['#2563eb', '#93c5fd'],
                    text=[f'{v:,.1f}' for v in compare_qty], textposition='outside',
                ), secondary_y=False)

                fig_annual.add_trace(go.Scatter(
                    x=compare_labels, y=compare_dot, name='DOT值',
                    mode='lines+markers+text', line=dict(color='#ea580c', width=2.5),
                    marker=dict(size=10), text=[f'{v:.2f}' for v in compare_dot], textposition='top center',
                ), secondary_y=True)

                fig_annual.update_layout(
                    height=400, hovermode='x unified',
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                    margin=dict(l=20, r=20, t=35, b=20),
                )
                fig_annual.update_yaxes(title_text="累计购药总支数", secondary_y=False)
                fig_annual.update_yaxes(title_text="DOT 值", secondary_y=True)
                st.plotly_chart(fig_annual, use_container_width=True)

            else:  # 滚动年度DOT对比
                st.markdown("""
                <div class="def-box" style="font-size:0.82rem;">
                <strong>滚动年度DOT对比：</strong>基于所选时间范围的最末次时间，向前取12个月作为滚动窗口，与上一个12个月窗口进行对比。<br/>
                例：选择 <strong>2026年1月~6月</strong>（最末次为2026年6月），则对比 <strong>2024.7~2025.6</strong> vs <strong>2025.7~2026.6</strong> 的数据。
                </div>
                """, unsafe_allow_html=True)

                col_c, col_d = st.columns(2)

                with col_c:
                    st.subheader(f"📅 当前滚动年：{ac['rolling']['current_label']}")
                    curr_r = ac['rolling']['current']
                    st.markdown(f"""
                    <div style="background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                        <p style="margin: 0.3rem 0;"><strong>去重患者数：</strong>{curr_r['patients']:,} 人</p>
                        <p style="margin: 0.3rem 0;"><strong>累计购药总支数：</strong>{curr_r['total_qty']:,.1f} 支</p>
                        <p style="margin: 0.3rem 0;"><strong>DOT 值：</strong><span style="font-size: 1.3rem; font-weight: 700; color: #2563eb;">{curr_r['dot']:.2f}</span></p>
                    </div>
                    """, unsafe_allow_html=True)

                with col_d:
                    st.subheader(f"📅 上一滚动年：{ac['rolling']['prev_label']}")
                    prev_r = ac['rolling']['prev']
                    if prev_r['patients'] > 0:
                        dot_change_r = ((curr_r['dot'] - prev_r['dot']) / prev_r['dot'] * 100) if prev_r['dot'] > 0 else 0
                        patient_change_r = ((curr_r['patients'] - prev_r['patients']) / prev_r['patients'] * 100) if prev_r['patients'] > 0 else 0
                        qty_change_r = ((curr_r['total_qty'] - prev_r['total_qty']) / prev_r['total_qty'] * 100) if prev_r['total_qty'] > 0 else 0

                        st.markdown(f"""
                        <div style="background: #fefce8; border: 1px solid #fde68a; border-radius: 8px; padding: 1rem; margin: 0.5rem 0;">
                            <p style="margin: 0.3rem 0;"><strong>去重患者数：</strong>{prev_r['patients']:,} 人
                            <span style="color: {'#16a34a' if patient_change_r >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if patient_change_r >= 0 else '↓'}{abs(patient_change_r):.1f}%)</span></p>
                            <p style="margin: 0.3rem 0;"><strong>累计购药总支数：</strong>{prev_r['total_qty']:,.1f} 支
                            <span style="color: {'#16a34a' if qty_change_r >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if qty_change_r >= 0 else '↓'}{abs(qty_change_r):.1f}%)</span></p>
                            <p style="margin: 0.3rem 0;"><strong>DOT 值：</strong><span style="font-size: 1.3rem; font-weight: 700; color: #ea580c;">{prev_r['dot']:.2f}</span>
                            <span style="color: {'#16a34a' if dot_change_r >= 0 else '#dc2626'}; font-size: 0.85rem;">({'↑' if dot_change_r >= 0 else '↓'}{abs(dot_change_r):.1f}%)</span></p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.info("上一滚动年无数据")

                # 滚动年度对比图表
                st.subheader("滚动年度DOT对比图表")

                rolling_labels = [ac['rolling']['current_label'], ac['rolling']['prev_label']]
                rolling_patients = [curr_r['patients'], prev_r['patients']]
                rolling_qty = [curr_r['total_qty'], prev_r['total_qty']]
                rolling_dot = [curr_r['dot'], prev_r['dot']]

                fig_rolling = make_subplots(specs=[[{"secondary_y": True}]])

                fig_rolling.add_trace(go.Bar(
                    x=rolling_labels, y=rolling_qty, name='累计购药总支数',
                    marker_color=['#16a34a', '#86efac'],
                    text=[f'{v:,.1f}' for v in rolling_qty], textposition='outside',
                ), secondary_y=False)

                fig_rolling.add_trace(go.Scatter(
                    x=rolling_labels, y=rolling_dot, name='DOT值',
                    mode='lines+markers+text', line=dict(color='#dc2626', width=2.5),
                    marker=dict(size=10), text=[f'{v:.2f}' for v in rolling_dot], textposition='top center',
                ), secondary_y=True)

                fig_rolling.update_layout(
                    height=400, hovermode='x unified',
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                    margin=dict(l=20, r=20, t=35, b=20),
                )
                fig_rolling.update_yaxes(title_text="累计购药总支数", secondary_y=False)
                fig_rolling.update_yaxes(title_text="DOT 值", secondary_y=True)
                st.plotly_chart(fig_rolling, use_container_width=True)

        # ==================== 脱落率分析 ====================
        st.markdown("---")
        st.header("📉 脱落率分析")

        dropout_data = ad.get('dropout_data', None)

        if dropout_data is None:
            st.info("未获取到脱落率数据，请重新点击「开始分析」。")
        elif not dropout_data.get('can_compute'):
            st.warning(f"⚠️ {dropout_data.get('reason', '无法计算脱落率')}")
        else:
            recent_label = "、".join(get_month_label(m) for m in dropout_data['recent_months'])
            prior_label = "、".join(get_month_label(m) for m in dropout_data['prior_months'])

            st.markdown(f"""
            <div class="def-box" style="font-size:0.82rem;">
            <strong>脱落率定义：</strong>在 <strong>{prior_label}</strong> 有购药记录的患者中，后续 <strong>{recent_label}</strong> 未再购药的患者占比。<br/>
            公式：脱落率 = 脱落患者数 / 倒推两个月前购药患者总人数 × 100%
            </div>
            """, unsafe_allow_html=True)

            # 指标卡片
            dc1, dc2, dc3 = st.columns(3)
            with dc1:
                st.markdown(f"""
                <div class="metric-card red">
                    <div class="label">脱落率</div>
                    <div class="value">{dropout_data['dropout_rate']:.2f}%</div>
                    <div class="sub">{prior_label} 购药但 {recent_label} 未购药</div>
                </div>
                """, unsafe_allow_html=True)
            with dc2:
                st.markdown(f"""
                <div class="metric-card orange">
                    <div class="label">脱落患者数</div>
                    <div class="value">{dropout_data['dropout_count']:,}</div>
                    <div class="sub">近两个月未购药</div>
                </div>
                """, unsafe_allow_html=True)
            with dc3:
                st.markdown(f"""
                <div class="metric-card blue">
                    <div class="label">倒推两个月前购药患者</div>
                    <div class="value">{dropout_data['denominator_count']:,}</div>
                    <div class="sub">作为分母</div>
                </div>
                """, unsafe_allow_html=True)

            # 名单 + 分布
            dcol1, dcol2 = st.columns(2)

            with dcol1:
                st.subheader("脱落患者名单")
                ddf = dropout_data['dropout_patients']
                if not ddf.empty:
                    st.dataframe(ddf, use_container_width=True, hide_index=True,
                                 column_config={'序号': st.column_config.NumberColumn(width='small'),
                                                '累计购药支数': st.column_config.NumberColumn(format='%.2f')})
                else:
                    st.info("没有符合脱落定义的患者")

            with dcol2:
                st.subheader("脱落患者购药支数分布")
                dist_df = dropout_data['dropout_distribution']
                if not dist_df.empty and dist_df['患者人数'].sum() > 0:
                    fig_dropout = go.Figure()
                    fig_dropout.add_trace(go.Bar(
                        y=list(dist_df['购药支数区间']), x=list(dist_df['患者人数']), orientation='h',
                        marker_color=['#2563eb', '#16a34a', '#ea580c', '#a855f7', '#ec4899', '#dc2626', '#64748b'],
                        text=[f"{c}人 ({r:.1f}%)" for c, r in zip(dist_df['患者人数'], dist_df['占比(%)'])],
                        textposition='outside',
                        hovertemplate='<b>%{y}</b><br>患者人数: %{x}<extra></extra>',
                    ))
                    fig_dropout.update_layout(
                        height=350, xaxis=dict(title="患者人数", rangemode='tozero'),
                        yaxis=dict(title=""), margin=dict(l=20, r=90, t=25, b=20), showlegend=False
                    )
                    st.plotly_chart(fig_dropout, use_container_width=True)
                else:
                    st.info("没有分布数据")

            # ---- 每月滚动脱落率趋势（新增脱落） ----
            st.subheader("每月新增脱落率趋势（滚动回顾）")
            st.caption("每月柱子/折线表示该月**新增**脱落人数（即患者首次满足脱落条件的那个月），不是持续未购药的累计人数。对回顾月份 M（基准月 T-2 = M 的前两个月）：分母 = 仅在基准月 T-2 有购药的患者；脱落 = 分母中在 T-1、T（即 M-1、M）都未购药的患者；脱落率 = 脱落 / 分母。时间段最早的两个月无法回顾，不显示。")
            trend_df = dropout_data.get('monthly_trend', pd.DataFrame())
            if trend_df is not None and not trend_df.empty:
                fig_trend = make_subplots(specs=[[{"secondary_y": True}]])
                fig_trend.add_trace(go.Bar(
                    x=list(trend_df['回顾月份']), y=list(trend_df['脱落人数']),
                    name='脱落人数', marker_color='#f97316',
                    text=list(trend_df['脱落人数']), textposition='outside',
                    hovertemplate='<b>%{x}</b><br>脱落人数: %{y}<extra></extra>',
                ), secondary_y=False)
                fig_trend.add_trace(go.Scatter(
                    x=list(trend_df['回顾月份']), y=list(trend_df['脱落率(%)']),
                    name='脱落率(%)', mode='lines+markers+text',
                    line=dict(color='#dc2626', width=3), marker=dict(size=8),
                    text=[f"{v:.1f}%" for v in trend_df['脱落率(%)']], textposition='top center',
                    hovertemplate='<b>%{x}</b><br>脱落率: %{y:.2f}%<extra></extra>',
                ), secondary_y=True)
                fig_trend.update_layout(
                    height=420, hovermode='x unified',
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                    margin=dict(l=20, r=20, t=40, b=20),
                )
                fig_trend.update_yaxes(title_text="脱落人数", secondary_y=False, rangemode='tozero')
                fig_trend.update_yaxes(title_text="脱落率 (%)", secondary_y=True, rangemode='tozero')
                st.plotly_chart(fig_trend, use_container_width=True)
                with st.expander("查看每月脱落率明细表", expanded=False):
                    st.dataframe(
                        trend_df[['回顾月份', '脱落人数', '分母人数', '脱落率(%)']],
                        use_container_width=True, hide_index=True,
                        column_config={'脱落率(%)': st.column_config.NumberColumn(format='%.2f')}
                    )
            else:
                st.info("分析时间段不足，无法生成每月脱落率趋势（至少需要 3 个月）。")

        # ==================== 图表区域 ====================
        st.markdown("---")
        st.header("📈 数据可视化")

        all_months_sorted = sorted(monthly_stats.keys())
        ml_all = [get_month_label(m) for m in all_months_sorted]
        qty_all = [round(monthly_stats[m]['qty'], 2) for m in all_months_sorted]
        active_pats_all = [len(monthly_stats[m]['patients']) for m in all_months_sorted]

        # 新患数据
        new_pats_all = [len(new_patient_by_month.get(m, set())) for m in all_months_sorted]

        # 高亮选中区间
        bar_colors = []
        for m in all_months_sorted:
            if result['start_month'] <= m <= result['end_month']:
                bar_colors.append('rgba(37, 99, 235, 0.78)')
            else:
                bar_colors.append('rgba(148, 163, 184, 0.35)')

        # --- 图表1: 月度趋势（活跃+新患+销量）---
        st.subheader("月度购药趋势（活跃患者数 + 新患患者数 + 购药支数）")

        fig_trend = make_subplots(specs=[[{"secondary_y": True}]])
        fig_trend.add_trace(go.Bar(
            x=ml_all, y=qty_all, name='购药总支数',
            marker_color=bar_colors,
            hovertemplate='<b>%{x}</b><br>购药总支数: %{y}<extra></extra>',
        ), secondary_y=False)
        fig_trend.add_trace(go.Scatter(
            x=ml_all, y=active_pats_all, name='活跃患者数',
            mode='lines+markers+text', line=dict(color='#2563eb', width=2.5),
            marker=dict(size=7), text=active_pats_all, textposition='top center',
            hovertemplate='<b>%{x}</b><br>活跃患者数: %{y}<extra></extra>',
        ), secondary_y=True)
        fig_trend.add_trace(go.Scatter(
            x=ml_all, y=new_pats_all, name='新患患者数',
            mode='lines+markers+text', line=dict(color='#dc2626', width=2, dash='dash'),
            marker=dict(size=6, symbol='diamond'), text=new_pats_all, textposition='bottom center',
            hovertemplate='<b>%{x}</b><br>新患患者数: %{y}<extra></extra>',
        ), secondary_y=True)
        fig_trend.update_layout(
            height=420, hovermode='x unified',
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            margin=dict(l=20, r=20, t=35, b=20),
        )
        fig_trend.update_yaxes(title_text="购药总支数", secondary_y=False)
        fig_trend.update_yaxes(title_text="患者人数", secondary_y=True)
        st.plotly_chart(fig_trend, use_container_width=True)

        # --- 图表2: 月度 DOT 对比（滚动一年算法）---
        st.subheader("月度 DOT 对比（滚动 12 个月窗口）")
        st.markdown("""
        <div class="def-box" style="font-size:0.82rem;">
        <strong>滚动一年算法：</strong>以目标月份为终点，向前取连续 12 个月作为窗口期计算 DOT。<br/>
        例：<strong>2026 年 1 月的 DOT</strong> = 2025 年 2 月 ~ 2026 年 1 月期间的「总购药支数 ÷ 去重患者数」
        </div>
        """, unsafe_allow_html=True)

        rolling_dot_data = []
        rolling_dot_labels = []
        for target_mk in all_months_sorted:
            window_start = target_mk
            # 向前推 11 个月
            ws = target_mk
            for _ in range(11):
                ws = prev_month(ws)
                if ws is None:
                    break
            if ws is None:
                continue

            window_months = []
            cur = ws
            for _ in range(12):
                if cur is None:
                    break
                window_months.append(cur)
                # 下一个月
                parts = cur.split('-')
                y, m = int(parts[0]), int(parts[1])
                m += 1
                if m > 12:
                    m = 1; y += 1
                cur = f"{y}-{m:02d}"

            window_patients = set()
            window_qty = 0
            for wm in window_months:
                if wm in monthly_stats:
                    window_patients.update(monthly_stats[wm]['patients'])
                    window_qty += monthly_stats[wm]['qty']

            wp_count = len(window_patients)
            rd = round(window_qty / wp_count, 2) if wp_count > 0 else 0
            rolling_dot_data.append(rd)
            rolling_dot_labels.append(get_month_label(target_mk))

        # 只展示用户选中的时间段（如果有的话），否则展示全部
        display_idx = [i for i, mk in enumerate(all_months_sorted)
                       if result['start_month'] <= mk <= result['end_month']]
        if not display_idx:
            display_idx = list(range(len(rolling_dot_labels)))

        display_labels = [rolling_dot_labels[i] for i in display_idx]
        display_dots = [rolling_dot_data[i] for i in display_idx]

        fig_dot = go.Figure()
        fig_dot.add_trace(go.Bar(
            x=display_labels, y=display_dots, name='月度 DOT（滚动12月）',
            marker_color='rgba(22, 163, 74, 0.72)',
            marker_line_color='rgba(22, 163, 74, 1)', marker_line_width=1,
            text=[f'{v:.2f}' for v in display_dots], textposition='outside',
            hovertemplate='<b>%{x}</b><br>滚动12月 DOT: %{y:.2f}<extra></extra>',
        ))
        # 添加平均线
        avg_dot = sum(display_dots) / len(display_dots) if display_dots else 0
        fig_dot.add_hline(y=avg_dot, line_dash="dot", line_color="#ef4444",
                          annotation_text=f"均值: {avg_dot:.2f}", annotation_position="top right")
        fig_dot.update_layout(
            height=380, yaxis=dict(title="DOT 值", rangemode='tozero'),
            xaxis=dict(title="", tickangle=-30),
            margin=dict(l=20, r=30, t=25, b=50), showlegend=False,
        )
        st.plotly_chart(fig_dot, use_container_width=True)

        # --- 图表3: 新患 vs 活跃对比 ---
        st.subheader("活跃患者数 vs 新患患者数（月度对比）")

        period_mks = [m for m in all_months_sorted if result['start_month'] <= m <= result['end_month']]
        period_labels = [get_month_label(m) for m in period_mks]
        active_in_period = [len(monthly_stats[m]['patients']) for m in period_mks]
        new_in_period = [len(new_patient_by_month.get(m, set())) for m in period_mks]

        fig_compare = go.Figure()
        fig_compare.add_trace(go.Bar(
            x=period_labels, y=active_in_period, name='活跃患者数',
            marker_color='rgba(37, 99, 235, 0.7)',
            text=active_in_period, textposition='outside',
        ))
        fig_compare.add_trace(go.Bar(
            x=period_labels, y=new_in_period, name='新患患者数',
            marker_color='rgba(220, 38, 38, 0.7)',
            text=new_in_period, textposition='outside',
        ))
        fig_compare.update_layout(
            height=350, barmode='group', yaxis=dict(title="人数", rangemode='tozero'),
            xaxis=dict(title="", tickangle=-30),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            margin=dict(l=20, r=20, t=25, b=55),
        )
        st.plotly_chart(fig_compare, use_container_width=True)

        # --- 图表4: 患者购药支数分布 ---
        st.subheader(f"患者购药支数分布（统计范围：{get_month_label(result['start_month'])} ~ {get_month_label(result['end_month'])}）")

        buckets = [('1支', 1, 1), ('2-3支', 2, 3), ('4-6支', 4, 6),
                     ('7-12支', 7, 12), ('13-24支', 13, 24), ('25-36支', 25, 36), ('37支以上', 37, float('inf'))]
        bucket_counts = []
        for label, lo, hi in buckets:
            cnt = len(details_df[(details_df['购药总支数'] >= lo) & (details_df['购药总支数'] <= hi)])
            bucket_counts.append(cnt)

        bl = [b[0] for b in buckets]
        dc = ['#2563eb', '#16a34a', '#ea580c', '#a855f7', '#ec4899', '#dc2626', '#64748b']

        fig_dist = go.Figure()
        fig_dist.add_trace(go.Bar(
            y=bl, x=bucket_counts, orientation='h', marker_color=dc,
            text=[f'{c}人 ({c/result["unique_patient_count"]*100:.1f}%)'
                  if result['unique_patient_count'] > 0 else f'{c}人' for c in bucket_counts],
            textposition='outside',
            hovertemplate='<b>%{y}</b><br>患者人数: %{x}<extra></extra>',
        ))
        fig_dist.update_layout(height=350, xaxis=dict(title="患者人数", rangemode='tozero'),
                               yaxis=dict(title=""), margin=dict(l=20, r=90, t=25, b=20), showlegend=False)
        st.plotly_chart(fig_dist, use_container_width=True)

        # ==================== 汇总表格区域 ====================
        st.markdown("---")
        tc1, tc2 = st.columns(2)

        with tc1:
            # 品种汇总
            psdf = ad.get('product_summary_df', pd.DataFrame())
            if not psdf.empty:
                st.subheader("💊 品种汇总数据")
                st.dataframe(psdf, use_container_width=True, hide_index=True,
                             column_config={'序号': st.column_config.NumberColumn(width='small'),
                                            '去重患者数': st.column_config.NumberColumn(format='%d'),
                                            '购药总次数': st.column_config.NumberColumn(format='%d'),
                                            '累计购药总支数': st.column_config.NumberColumn(format='%.2f'),
                                            'DOT值': st.column_config.NumberColumn(format='%.2f')})

        with tc2:
            # 药房汇总
            phsdf = ad.get('pharmacy_summary_df', pd.DataFrame())
            if not phsdf.empty:
                st.subheader("🏪 药房汇总数据")
                st.dataframe(phsdf, use_container_width=True, hide_index=True,
                             column_config={'序号': st.column_config.NumberColumn(width='small'),
                                            '去重患者数': st.column_config.NumberColumn(format='%d'),
                                            '购药总次数': st.column_config.NumberColumn(format='%d'),
                                            '累计购药总支数': st.column_config.NumberColumn(format='%.2f'),
                                            'DOT值': st.column_config.NumberColumn(format='%.2f')})

        # 品种×药房交叉汇总
        csdf = ad.get('cross_summary_df', pd.DataFrame())
        if not csdf.empty:
            st.subheader("📊 品种 × 药房 交叉汇总")
            st.dataframe(csdf, use_container_width=True, hide_index=True,
                         column_config={'序号': st.column_config.NumberColumn(width='small'),
                                        '去重患者数': st.column_config.NumberColumn(format='%d'),
                                        '购药总次数': st.column_config.NumberColumn(format='%d'),
                                        '累计购药总支数': st.column_config.NumberColumn(format='%.2f'),
                                        'DOT值': st.column_config.NumberColumn(format='%.2f')})

        # ==================== 患者明细表格 ====================
        st.markdown("---")
        st.header("📋 患者明细数据")

        page_size = st.selectbox("每页显示", [20, 50, 100], index=0, key='page_size')
        tp = max(1, (len(details_df) + page_size - 1) // page_size)
        cc1, cc2, cc3 = st.columns([1, 2, 1])
        with cc2:
            page = st.number_input("页码", min_value=1, max_value=tp, value=1, step=1, key='current_page')

        si_ = (page - 1) * page_size
        ei_ = min(si_ + page_size, len(details_df))
        st.dataframe(details_df.iloc[si_:ei_], use_container_width=True, hide_index=True,
                      column_config={'序号': st.column_config.NumberColumn(width='small'),
                                     '购药总支数': st.column_config.NumberColumn(format='%.2f'),
                                     '购药总次数': st.column_config.NumberColumn(format='%d')})
        st.caption(f"显示第 {si_ + 1} ~ {ei_} 条，共 {len(details_df)} 条记录 · 第 {page}/{tp} 页")

        # ==================== 导出 ====================
        st.markdown("---")
        st.header("📥 数据下载")

        st.markdown(f"""
        <div class="def-box">
        <strong>📥 第一大按钮：导出完整分析报告</strong><br/><br/>
        <strong>📋 Sheet 1 — 汇总：</strong>原始底表数据 + 左右并排的药房DOT表和品种×药房透视汇总<br/><br/>
        <strong>📊 Sheet 2-12 — 全局汇总维度：</strong><br/>
        &nbsp;&nbsp;• 分析摘要、月度趋势、月度DOT对比、活跃vs新患、患者分布、年度DOT对比、<strong>脱落率分析</strong>、品种/药房/患者明细<br/>
        &nbsp;&nbsp;• <em>每个Sheet标题行上方标注数据计算时间段（格式：YYYY年MM月DD日～YYYY年MM月DD日）</em><br/><br/>
        <strong>🏪 Sheet 13+ — 分药房维度（合并版）：</strong><br/>
        &nbsp;&nbsp;• 每个药房一个Sheet，包含5大分析维度（月度趋势 / DOT对比 / 活跃vs新患 / 患者分布 / 年度DOT）<br/>
        &nbsp;&nbsp;• 各维度以<span style='background:#FFD700;padding:0 4px;'>黄色标题</span>分隔，视觉清晰，Sheet数量大幅减少<br/>
        &nbsp;&nbsp;• <em>同一药房所有数据100%在同一Sheet中，杜绝拆分</em><br/><br/>
        <strong>📊 第二大按钮：单独下载透视汇总表</strong><br/>
        &nbsp;&nbsp;• 独立文件，包含品种×药房透视汇总、药房DOT汇总、品种汇总三个Sheet<br/>
        &nbsp;&nbsp;• 适合只需要对照分析数据的使用场景
        </div>
        """, unsafe_allow_html=True)

        excel_buf = BytesIO()
        with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
            # 【防御·关键】确保工作簿至少有一个【可见】Sheet（覆盖「空工作簿」与「全部隐藏」两种情况）。
            # 部分运行环境（openpyxl 在某些 pandas 版本下）创建工作簿时不带默认 Sheet；
            # 若导出逻辑在创建任何 Sheet 之前抛出早期异常，save() 会先抛
            # IndexError("At least one sheet must be visible") 把真实错误掩盖。
            # 预先保证有可见 Sheet，即使后续出错 save() 也能成功，从而让【真实错误】冒泡出来。
            _wbs = getattr(writer.book, 'worksheets', [])
            if not any(getattr(ws, 'sheet_state', 'visible') == 'visible' for ws in _wbs):
                if _wbs:
                    _wbs[0].sheet_state = 'visible'
                else:
                    writer.book.create_sheet('Sheet1')
            from openpyxl.styles import Font, PatternFill, Alignment

            sm = result.get('start_month'); em = result.get('end_month')
            if not sm or not em:
                st.error("⚠️ 当前分析结果缺少起止月份（通常因未识别到有效日期列）。请检查数据中的日期列后，重新运行分析，再导出 Excel。")
                st.stop()
            ad_local = st.session_state.get('analysis_data') or {}
            ac_local = st.session_state.get('annual_comparison') or {}
            date_col_local = ad_local.get('date_col', None)
            if not date_col_local:
                st.error("⚠️ 未识别到『购药时间』日期列，无法生成带日期范围的 Excel。请返回上方『字段配置』选择『购药时间 *』后，重新运行分析，再导出。")
                st.stop()

            # ---- 计算实际日期范围 ----
            filtered_df_export2 = ad_local.get('filtered_df', pd.DataFrame())

            def _fmt_date(d):
                """将日期转为 YYYY年MM月DD日（对 NaT/非法日期容错）"""
                if d is None or (isinstance(d, float) and pd.isna(d)):
                    return '未知'
                if isinstance(d, pd.Timestamp):
                    if d is pd.NaT:
                        return '未知'
                    return f"{d.year}年{d.month}月{d.day}日"
                if hasattr(d, 'year'):
                    try:
                        return f"{d.year}年{d.month}月{d.day}日"
                    except Exception:
                        return '未知'
                return str(d)

            if date_col_local and not filtered_df_export2.empty and date_col_local in filtered_df_export2.columns:
                try:
                    # 关键：先把日期列安全转为 datetime64 再取最值。
                    # 否则当该列为 object（字符串/混合类型，如部分 Timestamp + 部分 None/字符串）
                    # 时，Series.min() 会抛 TypeError（umr_minimum 无法比较）。
                    _dt_series = pd.to_datetime(filtered_df_export2[date_col_local], errors='coerce')
                    start_dt = _dt_series.min()
                    end_dt = _dt_series.max()
                except Exception:
                    start_dt = end_dt = None
                if start_dt is None or end_dt is None or pd.isna(start_dt) or pd.isna(end_dt):
                    start_dt = end_dt = None  # 触发下方回退（按月份key推算）
            else:
                start_dt = end_dt = None
            if start_dt is None or end_dt is None:
                # 回退：从月份key推算
                sp = sm.split('-'); ep = em.split('-')
                sy, sm2 = int(sp[0]), int(sp[1])
                ey, em2 = int(ep[0]), int(ep[1])
                import calendar as _cal
                last_day = _cal.monthrange(ey, em2)[1]
                start_dt = f"{sy}年{sm2}月1日"
                end_dt = f"{ey}年{em2}月{last_day}日"

            time_label = f"数据计算时间段：{_fmt_date(start_dt)} ～ {_fmt_date(end_dt)}"

            # ---- 通用样式 ----
            section_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            section_font = Font(bold=True, size=11, color="333333")
            time_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            time_font = Font(bold=True, color="0B5394", size=11)
            header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
            header_font = Font(bold=True, size=10)

            def add_time_header(ws, max_col=None):
                """在Sheet顶部插入时间范围标注行"""
                if max_col is None:
                    max_col = ws.max_column
                ws.insert_rows(1)
                if max_col > 1:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                cell = ws.cell(row=1, column=1, value=time_label)
                cell.font = time_font
                cell.fill = time_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

            def write_section_title(ws, row, col, title, end_col):
                """写入黄色背景维度标题"""
                if end_col > col:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
                cell = ws.cell(row=row, column=col, value=title)
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                return row + 1

            # ============================================================
            # Sheet 1 (NEW): 汇总总表
            # ============================================================
            filtered_df_export = ad_local.get('filtered_df', pd.DataFrame())
            phsdf_export = ad_local.get('pharmacy_summary_df', pd.DataFrame())
            csdf_export = ad_local.get('cross_summary_df', pd.DataFrame())

            ws_master = writer.book.create_sheet('汇总', 0)

            # --- 总标题 ---
            ws_master.merge_cells('A1:N1')
            c = ws_master.cell(row=1, column=1, value='阿斯利康患者数据分析 — 汇总')
            c.font = Font(bold=True, size=16, color='1a3a5c')
            c.alignment = Alignment(horizontal='center', vertical='center')

            ws_master.merge_cells('A2:N2')
            c = ws_master.cell(row=2, column=1, value=time_label)
            c.font = time_font; c.fill = time_fill
            c.alignment = Alignment(horizontal='left', vertical='center')

            mr = 4

            # --- 第一部分：原始底表数据 ---
            write_section_title(ws_master, mr, 1, '第一部分：原始底表数据', 14); mr += 1

            if not filtered_df_export.empty:
                key_cols = list(filtered_df_export.columns[:min(15, len(filtered_df_export.columns))])
                sub_df = filtered_df_export[key_cols].head(500)

                for ci, col_name in enumerate(key_cols, 1):
                    c = ws_master.cell(row=mr, column=ci, value=col_name)
                    c.font = header_font; c.fill = header_fill
                mr += 1

                for _, row in sub_df.iterrows():
                    for ci, col_name in enumerate(key_cols, 1):
                        val = row[col_name]
                        if pd.isna(val): val = ''
                        ws_master.cell(row=mr, column=ci, value=val)
                    mr += 1

                if len(filtered_df_export) > 500:
                    ws_master.merge_cells(start_row=mr, start_column=1, end_row=mr, end_column=len(key_cols))
                    c = ws_master.cell(row=mr, column=1,
                        value=f'注：原始数据共 {len(filtered_df_export):,} 行，此处仅展示前 500 行预览')
                    c.font = Font(italic=True, color='888888', size=10)
                    mr += 1
            else:
                ws_master.cell(row=mr, column=1, value='（无原始底表数据）'); mr += 1

            mr += 2

            # --- 第二部分：计算结果 ---
            write_section_title(ws_master, mr, 1, '第二部分：计算结果 — 左侧：各药房DOT数据  |  右侧：品种×药房透视汇总', 14)
            mr += 2

            # ------ 左侧：药房DOT数据表（从 A 列开始）------
            left_top = mr
            if not phsdf_export.empty:
                c = ws_master.cell(row=mr, column=1, value='各药房 DOT 数据表')
                c.font = Font(bold=True, size=11, color='1a3a5c'); mr += 1

                ph_cols = list(phsdf_export.columns)
                for ci, cn in enumerate(ph_cols, 1):
                    c = ws_master.cell(row=mr, column=ci, value=cn)
                    c.font = header_font
                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                mr += 1

                for _, row in phsdf_export.iterrows():
                    for ci, cn in enumerate(ph_cols, 1):
                        ws_master.cell(row=mr, column=ci, value=row[cn])
                    mr += 1
            else:
                ws_master.cell(row=mr, column=1, value='（无药房数据）'); mr += 1
            left_bottom = mr - 1

            # ------ 右侧：品种×药房透视汇总（从 H 列开始）------
            right_col = 8
            right_top = left_top
            mr_right = left_top

            if not csdf_export.empty:
                c = ws_master.cell(row=mr_right, column=right_col, value='品种×药房 透视汇总表')
                c.font = Font(bold=True, size=11, color='1a3a5c'); mr_right += 1

                cs_cols = list(csdf_export.columns)
                for ci, cn in enumerate(cs_cols):
                    c = ws_master.cell(row=mr_right, column=right_col + ci, value=cn)
                    c.font = header_font
                    c.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                mr_right += 1

                for _, row in csdf_export.iterrows():
                    for ci, cn in enumerate(cs_cols):
                        ws_master.cell(row=mr_right, column=right_col + ci, value=row[cn])
                    mr_right += 1
            else:
                c = ws_master.cell(row=mr_right, column=right_col, value='品种×药房 透视汇总表')
                c.font = Font(bold=True, size=11, color='1a3a5c'); mr_right += 1
                ws_master.cell(row=mr_right, column=right_col, value='（需要同时配置品种和药房字段）'); mr_right += 1

            # 设置列宽
            width_map = {'A': 8, 'B': 22, 'C': 14, 'D': 14, 'E': 18, 'F': 12, 'G': 12,
                         'H': 10, 'I': 20, 'J': 22, 'K': 14, 'L': 14, 'M': 18, 'N': 12}
            for col_letter, w in width_map.items():
                ws_master.column_dimensions[col_letter].width = w

            # ============================================================
            # Sheet 2: 分析摘要
            # ============================================================
            summary_info = [
                ['阿斯利康患者数据分析报告', ''],
                ['', ''],
                ['一、筛选条件', ''],
                ['时间范围', f'{get_month_label(sm)} ~ {get_month_label(em)}'],
                ['药房筛选', ', '.join(result.get('selected_pharmacies', ['全部']))],
                ['品种筛选', ', '.join(result.get('selected_products', ['全部']))],
                ['', ''],
                ['二、核心指标', '数值'],
                ['活跃患者数（定义：时间段内有≥1次购药记录的去重患者数）', result['unique_patient_count']],
                ['新患患者数（时间段内首次购药患者总数）',
                 sum(len(new_patient_by_month.get(m, set())) for m in sorted(monthly_stats.keys()) if sm <= m <= em)],
                ['累计购药总支数（活跃患者从首次到末次的全部购药）', result['total_quantity']],
                ['DOT 值（= 总支数 / 活跃患者数）', result['dot_value']],
                ['脱落率(%)（在「倒推两个月前」购药但「近两月」未购药的患者占比）',
                 ad_local.get('dropout_data', {}).get('dropout_rate') if ad_local.get('dropout_data', {}).get('can_compute') else '无法计算'],
                ['脱落患者数',
                 ad_local.get('dropout_data', {}).get('dropout_count') if ad_local.get('dropout_data', {}).get('can_compute') else '无法计算'],
                ['倒推两个月前购药患者数（分母）',
                 ad_local.get('dropout_data', {}).get('denominator_count') if ad_local.get('dropout_data', {}).get('can_compute') else '无法计算'],
                ['', ''],
                ['三、指标说明', ''],
                ['活跃患者数', '所选时间段内至少有 1 次购药记录的去重患者人数。统计范围为这些患者从首次购药到最新数据的全部记录。'],
                ['新患患者数', '在该月份购药且为全量数据中首次出现（即第一次购买）的去重患者人数。'],
                ['DOT 值', 'Duration of Treatment，用药时长指标。DOT = 累计购药总支数 / 去重患者数。'],
                ['滚动一年算法', '以目标月份为终点，向前取连续 12 个月作为窗口期计算 DOT。例：2026年1月DOT = 2025年2月至2026年1月的数据。'],
                ['', ''],
                ['四、年度DOT对比说明', ''],
                ['同期年度DOT对比', '将所选时间段与去年同期的DOT数据进行对比。例：选择2026年1月~6月，则对比2025年1月~6月的数据。'],
                ['滚动年度DOT对比', '基于所选时间范围的最末次时间，向前取12个月作为滚动窗口，与上一个12个月窗口进行对比。'],
            ]
            pd.DataFrame(summary_info, columns=['项目', '内容']).to_excel(writer, sheet_name='分析摘要', index=False)
            ws_sum = writer.sheets['分析摘要']
            ws_sum.column_dimensions['A'].width = 45
            ws_sum.column_dimensions['B'].width = 70
            add_time_header(ws_sum, 2)

            # ============================================================
            # Sheet 3: 月度趋势数据
            # ============================================================
            all_months_sorted = ad_local.get('all_months_sorted', [])
            ml_all = ad_local.get('ml_all', [])
            qty_all = ad_local.get('qty_all', [])
            active_pats_all = ad_local.get('active_pats_all', [])
            new_pats_all = ad_local.get('new_pats_all', [])

            trend_data = []
            for i, mk in enumerate(all_months_sorted):
                trend_data.append({
                    '月份': ml_all[i] if i < len(ml_all) else mk,
                    '购药总支数': qty_all[i] if i < len(qty_all) else 0,
                    '活跃患者数': active_pats_all[i] if i < len(active_pats_all) else 0,
                    '新患患者数': new_pats_all[i] if i < len(new_pats_all) else 0,
                })
            trend_df = pd.DataFrame(trend_data)
            trend_df.to_excel(writer, sheet_name='月度趋势数据', index=False)
            ws2 = writer.sheets['月度趋势数据']
            for ci, w in enumerate([15, 15, 15, 15], 1):
                ws2.column_dimensions[chr(64+ci)].width = w
            add_time_header(ws2, 4)

            # ============================================================
            # Sheet 4: 月度DOT对比数据
            # ============================================================
            rolling_dot_data = []
            rolling_dot_labels = []
            monthly_stats_local = ad_local.get('monthly_stats', {})

            for target_mk in all_months_sorted:
                ws = target_mk
                for _ in range(11):
                    ws = prev_month(ws)
                    if ws is None: break
                if ws is None: continue

                window_months = []
                cur = ws
                for _ in range(12):
                    if cur is None: break
                    window_months.append(cur)
                    parts = cur.split('-')
                    y, m = int(parts[0]), int(parts[1])
                    m += 1
                    if m > 12: m = 1; y += 1
                    cur = f"{y}-{m:02d}"

                window_patients = set()
                window_qty = 0
                for wm in window_months:
                    if wm in monthly_stats_local:
                        window_patients.update(monthly_stats_local[wm]['patients'])
                        window_qty += monthly_stats_local[wm]['qty']

                wp_count = len(window_patients)
                rd = round(window_qty / wp_count, 2) if wp_count > 0 else 0
                rolling_dot_data.append({
                    '月份': get_month_label(target_mk),
                    '滚动12个月DOT值': rd,
                    '窗口期开始': get_month_label(window_months[0]) if window_months else '-',
                    '窗口期结束': get_month_label(window_months[-1]) if window_months else '-',
                    '窗口期去重患者数': wp_count,
                    '窗口期累计购药支数': round(window_qty, 2),
                })
                rolling_dot_labels.append(get_month_label(target_mk))

            rolling_dot_df = pd.DataFrame(rolling_dot_data)
            rolling_dot_df.to_excel(writer, sheet_name='月度DOT对比数据', index=False)
            ws3 = writer.sheets['月度DOT对比数据']
            for ci, w in enumerate([15, 18, 15, 15, 18, 18], 1):
                ws3.column_dimensions[chr(64+ci)].width = w
            add_time_header(ws3, 6)

            # ============================================================
            # Sheet 5: 活跃vs新患对比数据
            # ============================================================
            compare_data = []
            for i, mk in enumerate(all_months_sorted):
                if sm <= mk <= em:
                    compare_data.append({
                        '月份': ml_all[i] if i < len(ml_all) else mk,
                        '活跃患者数': active_pats_all[i] if i < len(active_pats_all) else 0,
                        '新患患者数': new_pats_all[i] if i < len(new_pats_all) else 0,
                        '新患占比(%)': round((new_pats_all[i] / active_pats_all[i] * 100), 2) if i < len(active_pats_all) and active_pats_all[i] > 0 else 0,
                    })
            compare_df = pd.DataFrame(compare_data)
            compare_df.to_excel(writer, sheet_name='活跃vs新患对比', index=False)
            ws4 = writer.sheets['活跃vs新患对比']
            for ci, w in enumerate([15, 15, 15, 15], 1):
                ws4.column_dimensions[chr(64+ci)].width = w
            add_time_header(ws4, 4)

            # ============================================================
            # Sheet 6: 患者分布数据
            # ============================================================
            buckets = [('1支', 1, 1), ('2-3支', 2, 3), ('4-6支', 4, 6),
                       ('7-12支', 7, 12), ('13-24支', 13, 24), ('25-36支', 25, 36), ('37支以上', 37, float('inf'))]
            bucket_data = []
            for label, lo, hi in buckets:
                cnt = len(details_df[(details_df['购药总支数'] >= lo) & (details_df['购药总支数'] <= hi)])
                bucket_data.append({
                    '购药支数区间': label,
                    '患者人数': cnt,
                    '占比(%)': round(cnt / result['unique_patient_count'] * 100, 2) if result['unique_patient_count'] > 0 else 0,
                })
            bucket_df = pd.DataFrame(bucket_data)
            bucket_df.to_excel(writer, sheet_name='患者分布数据', index=False)
            ws5 = writer.sheets['患者分布数据']
            for ci, w in enumerate([15, 15, 15], 1):
                ws5.column_dimensions[chr(64+ci)].width = w
            add_time_header(ws5, 3)

            # ============================================================
            # Sheet 7: 年度DOT对比数据
            # ============================================================
            if ac_local is not None:
                annual_data = []
                same_curr = ac_local['same_period']['current']
                same_last = ac_local['same_period']['last']
                annual_data.append({
                    '对比类型': '同期年度DOT对比',
                    '当前期间': ac_local['same_period']['current_label'],
                    '去年同期': ac_local['same_period']['last_label'],
                    '当前去重患者数': same_curr['patients'],
                    '去年去重患者数': same_last['patients'],
                    '当前累计购药支数': same_curr['total_qty'],
                    '去年累计购药支数': same_last['total_qty'],
                    '当前DOT值': same_curr['dot'],
                    '去年DOT值': same_last['dot'],
                    'DOT同比变化(%)': round(((same_curr['dot'] - same_last['dot']) / same_last['dot'] * 100), 2) if same_last['dot'] > 0 else 0,
                })
                roll_curr = ac_local['rolling']['current']
                roll_prev = ac_local['rolling']['prev']
                annual_data.append({
                    '对比类型': '滚动年度DOT对比',
                    '当前期间': ac_local['rolling']['current_label'],
                    '上期期间': ac_local['rolling']['prev_label'],
                    '当前去重患者数': roll_curr['patients'],
                    '上期去重患者数': roll_prev['patients'],
                    '当前累计购药支数': roll_curr['total_qty'],
                    '上期累计购药支数': roll_prev['total_qty'],
                    '当前DOT值': roll_curr['dot'],
                    '上期DOT值': roll_prev['dot'],
                    'DOT变化(%)': round(((roll_curr['dot'] - roll_prev['dot']) / roll_prev['dot'] * 100), 2) if roll_prev['dot'] > 0 else 0,
                })
                annual_df = pd.DataFrame(annual_data)
                annual_df.to_excel(writer, sheet_name='年度DOT对比', index=False)
                ws6 = writer.sheets['年度DOT对比']
                for ci, w in enumerate([20, 20, 20, 15, 15, 18, 18, 12, 12, 15], 1):
                    ws6.column_dimensions[chr(64+ci)].width = w
                add_time_header(ws6, 10)

            # ============================================================
            # Sheet 8: 脱落率分析
            # ============================================================
            dropout_data_local = ad_local.get('dropout_data', None)
            if dropout_data_local and dropout_data_local.get('can_compute'):
                recent_label_local = "、".join(get_month_label(m) for m in dropout_data_local['recent_months'])
                prior_label_local = "、".join(get_month_label(m) for m in dropout_data_local['prior_months'])
                dropout_summary_rows = [
                    ['项目', '数值'],
                    ['脱落率定义', f'在 {prior_label_local} 有购药记录的患者中，后续 {recent_label_local} 未再购药的患者占比'],
                    ['脱落率(%)', dropout_data_local['dropout_rate']],
                    ['脱落患者数', dropout_data_local['dropout_count']],
                    ['倒推两个月前购药患者数（分母）', dropout_data_local['denominator_count']],
                    ['近两月份', recent_label_local],
                    ['倒推两个月前月份', prior_label_local],
                ]
                dropout_summary_df = pd.DataFrame(dropout_summary_rows[1:], columns=dropout_summary_rows[0])
                dropout_summary_df.to_excel(writer, sheet_name='脱落率分析', index=False)
                ws_dropout = writer.sheets['脱落率分析']
                ws_dropout.column_dimensions['A'].width = 45
                ws_dropout.column_dimensions['B'].width = 25
                add_time_header(ws_dropout, 2)

                # 脱落患者名单
                ddp = dropout_data_local['dropout_patients']
                if not ddp.empty:
                    start_row = 9
                    c = ws_dropout.cell(row=start_row, column=1, value='脱落患者名单')
                    c.font = Font(bold=True, size=11, color='1a3a5c')
                    start_row += 1
                    for ci, cn in enumerate(ddp.columns, 1):
                        c = ws_dropout.cell(row=start_row, column=ci, value=cn)
                        c.font = header_font
                        c.fill = header_fill
                    start_row += 1
                    for _, row in ddp.iterrows():
                        for ci, cn in enumerate(ddp.columns, 1):
                            ws_dropout.cell(row=start_row, column=ci, value=row[cn])
                        start_row += 1
                    start_row += 1

                    # 脱落患者支数分布
                    ddd = dropout_data_local['dropout_distribution']
                    if not ddd.empty:
                        c = ws_dropout.cell(row=start_row, column=1, value='脱落患者购药支数分布')
                        c.font = Font(bold=True, size=11, color='1a3a5c')
                        start_row += 1
                        for ci, cn in enumerate(ddd.columns, 1):
                            c = ws_dropout.cell(row=start_row, column=ci, value=cn)
                            c.font = header_font
                            c.fill = header_fill
                        start_row += 1
                        for _, row in ddd.iterrows():
                            for ci, cn in enumerate(ddd.columns, 1):
                                ws_dropout.cell(row=start_row, column=ci, value=row[cn])
                            start_row += 1
                        start_row += 1

                    # 每月滚动脱落率趋势
                    mtrend = dropout_data_local.get('monthly_trend', pd.DataFrame())
                    if mtrend is not None and not mtrend.empty:
                        mtrend_out = mtrend[['回顾月份', '脱落人数', '分母人数', '脱落率(%)']]
                        c = ws_dropout.cell(row=start_row, column=1, value='每月新增脱落率趋势（滚动回顾）')
                        c.font = Font(bold=True, size=11, color='1a3a5c')
                        start_row += 1
                        for ci, cn in enumerate(mtrend_out.columns, 1):
                            c = ws_dropout.cell(row=start_row, column=ci, value=cn)
                            c.font = header_font
                            c.fill = header_fill
                        start_row += 1
                        for _, row in mtrend_out.iterrows():
                            for ci, cn in enumerate(mtrend_out.columns, 1):
                                ws_dropout.cell(row=start_row, column=ci, value=row[cn])
                            start_row += 1
            else:
                reason = (dropout_data_local.get('reason') if dropout_data_local else '无法计算') or '无法计算'
                pd.DataFrame({'提示': [reason]}).to_excel(writer, sheet_name='脱落率分析', index=False)
                ws_dropout = writer.sheets['脱落率分析']
                ws_dropout.column_dimensions['A'].width = 60
                add_time_header(ws_dropout, 1)

            # ============================================================
            # Sheet 9: 品种汇总
            # ============================================================
            psdf_export2 = ad_local.get('product_summary_df', pd.DataFrame())
            if not psdf_export2.empty:
                psdf_export2.to_excel(writer, sheet_name='品种汇总', index=False)
                ws7 = writer.sheets['品种汇总']
                for ci, w in enumerate([8, 25, 14, 14, 18, 10], 1):
                    ws7.column_dimensions[chr(64+ci)].width = w
                add_time_header(ws7, 6)
            else:
                pd.DataFrame({'提示': ['未检测到品种字段或无数据']}).to_excel(writer, sheet_name='品种汇总', index=False)

            # ============================================================
            # Sheet 10: 药房汇总
            # ============================================================
            phsdf_export2 = ad_local.get('pharmacy_summary_df', pd.DataFrame())
            if not phsdf_export2.empty:
                phsdf_export2.to_excel(writer, sheet_name='药房汇总', index=False)
                ws8 = writer.sheets['药房汇总']
                for ci, w in enumerate([8, 25, 14, 14, 18, 10], 1):
                    ws8.column_dimensions[chr(64+ci)].width = w
                add_time_header(ws8, 6)
            else:
                pd.DataFrame({'提示': ['未检测到药房字段或无数据']}).to_excel(writer, sheet_name='药房汇总', index=False)

            # ============================================================
            # Sheet 11: 品种×药房交叉汇总
            # ============================================================
            csdf_export2 = ad_local.get('cross_summary_df', pd.DataFrame())
            if not csdf_export2.empty:
                csdf_export2.to_excel(writer, sheet_name='品种×药房汇总', index=False)
                ws9 = writer.sheets['品种×药房汇总']
                for ci, w in enumerate([8, 22, 25, 14, 14, 18, 10], 1):
                    ws9.column_dimensions[chr(64+ci)].width = w
                add_time_header(ws9, 7)
            else:
                pd.DataFrame({'提示': ['需要同时配置品种和药房字段']}).to_excel(writer, sheet_name='品种×药房汇总', index=False)

            # ============================================================
            # Sheet 12: 患者明细
            # ============================================================
            details_df.to_excel(writer, sheet_name='患者明细', index=False)
            ws10 = writer.sheets['患者明细']
            for ci, w in enumerate([8, 15, 12, 8, 25, 16, 16, 12, 14], 1):
                ws10.column_dimensions[chr(64+ci)].width = w
            add_time_header(ws10, 9)

            # ============================================================
            # 第二部分：分药房维度（合并版：每药房1个Sheet）
            # ============================================================
            pharm_monthly_data = ad_local.get('pharmacy_monthly_data', {})
            pharm_new_pat_month = ad_local.get('pharmacy_new_patient_month', {})
            patient_data_local = ad_local.get('patient_data', {})
            sm_store = ad_local.get('start_month', sm)
            em_store = ad_local.get('end_month', em)

            buckets_def = [('1支', 1, 1), ('2-3支', 2, 3), ('4-6支', 4, 6),
                           ('7-12支', 7, 12), ('13-24支', 13, 24), ('25-36支', 25, 36), ('37支以上', 37, float('inf'))]

            # 辅助：为单个药房计算滚动DOT数据
            def calc_rolling_dot_for_pharm(p_monthly):
                p_months_sorted = sorted(p_monthly.keys())
                rows = []
                for target_mk in p_months_sorted:
                    ws_mk = target_mk
                    for _ in range(11):
                        ws_mk = prev_month(ws_mk)
                        if ws_mk is None: break
                    if ws_mk is None: continue
                    window_months = []
                    cur = ws_mk
                    for _ in range(12):
                        if cur is None: break
                        window_months.append(cur)
                        yy, mm = int(cur.split('-')[0]), int(cur.split('-')[1])
                        mm += 1
                        if mm > 12: mm = 1; yy += 1
                        cur = f"{yy}-{mm:02d}"
                    w_pats = set()
                    w_qty = 0
                    for wm in window_months:
                        if wm in p_monthly:
                            w_pats.update(p_monthly[wm]['patients'])
                            w_qty += p_monthly[wm]['qty']
                    wp = len(w_pats)
                    rows.append({
                        '月份': get_month_label(target_mk),
                        '滚动12个月DOT值': round(w_qty / wp, 2) if wp > 0 else 0,
                        '窗口期开始': get_month_label(window_months[0]) if window_months else '-',
                        '窗口期结束': get_month_label(window_months[-1]) if window_months else '-',
                        '窗口期去重患者数': wp,
                        '窗口期累计购药支数': round(w_qty, 2),
                    })
                return rows

            def calc_annual_dot_for_pharm(p_monthly, sm_local, em_local):
                def period_dot(p_m, start_mk, end_mk):
                    pats = set(); qty = 0
                    for mk, data in p_m.items():
                        if start_mk <= mk <= end_mk:
                            pats.update(data['patients']); qty += data['qty']
                    pc = len(pats)
                    return {'patients': pc, 'total_qty': round(qty, 2),
                            'dot': round(qty / pc, 2) if pc > 0 else 0}
                sy, ey = int(sm_local.split('-')[0]), int(em_local.split('-')[1])
                last_sm = f"{sy-1}-{sm_local.split('-')[1]}"
                last_em = f"{ey-1}-{em_local.split('-')[1]}"
                same_cur = period_dot(p_monthly, sm_local, em_local)
                same_last = period_dot(p_monthly, last_sm, last_em)
                roll_cur_end = em_local
                roll_cur_start = roll_cur_end
                for _ in range(11):
                    roll_cur_start = prev_month(roll_cur_start)
                    if roll_cur_start is None: break
                roll_prev_end = prev_month(roll_cur_start) if roll_cur_start else None
                roll_prev_start = roll_prev_end
                if roll_prev_end:
                    for _ in range(11):
                        roll_prev_start = prev_month(roll_prev_start)
                        if roll_prev_start is None: break
                roll_cur = period_dot(p_monthly, roll_cur_start or sm_local, roll_cur_end)
                roll_prev = period_dot(p_monthly, roll_prev_start or sm_local, roll_prev_end or em_local) if roll_prev_end else {'patients': 0, 'total_qty': 0, 'dot': 0}
                rows = []
                rows.append({
                    '对比类型': '同期年度DOT对比',
                    '当前期间': f"{get_month_label(sm_local)}~{get_month_label(em_local)}",
                    '去年同期': f"{get_month_label(last_sm)}~{get_month_label(last_em)}",
                    '当前去重患者数': same_cur['patients'], '去年去重患者数': same_last['patients'],
                    '当前累计购药支数': same_cur['total_qty'], '去年累计购药支数': same_last['total_qty'],
                    '当前DOT值': same_cur['dot'], '去年DOT值': same_last['dot'],
                    'DOT同比变化(%)': round((same_cur['dot'] - same_last['dot']) / same_last['dot'] * 100, 2) if same_last['dot'] > 0 else 0,
                })
                rows.append({
                    '对比类型': '滚动年度DOT对比',
                    '当前期间': f"{get_month_label(roll_cur_start or sm_local)}~{get_month_label(roll_cur_end)}",
                    '上期期间': f"{get_month_label(roll_prev_start or sm_local)}~{get_month_label(roll_prev_end or em_local)}" if roll_prev_end else '-',
                    '当前去重患者数': roll_cur['patients'], '上期去重患者数': roll_prev['patients'],
                    '当前累计购药支数': roll_cur['total_qty'], '上期累计购药支数': roll_prev['total_qty'],
                    '当前DOT值': roll_cur['dot'], '上期DOT值': roll_prev['dot'],
                    'DOT变化(%)': round((roll_cur['dot'] - roll_prev['dot']) / roll_prev['dot'] * 100, 2) if roll_prev['dot'] > 0 else 0,
                })
                return rows

            # 工作表名清洗：Excel 不允许 \ / * ? : [ ] 等字符，且最长 31 字符
            import re as _re_sheet
            _ILLEGAL_SHEET_CHARS = _re_sheet.compile(r'[\\/*?:\[\]]')

            def _safe_sheet_name(name, maxlen=31):
                """去除 Excel 工作表名非法字符并截断到 maxlen（≤31）字符"""
                cleaned = _ILLEGAL_SHEET_CHARS.sub('', str(name)).strip()
                return (cleaned if cleaned else '药房')[:maxlen]

            pharm_list = sorted(pharm_monthly_data.keys())
            # 记录已用Sheet名，防止覆盖/拆分
            used_sheet_names = set()
            # 记录所有已存在的Sheet名（包括前面全局汇总创建的）
            for _s in writer.sheets:
                used_sheet_names.add(_s)

            for ph_name in pharm_list:
                p_m = pharm_monthly_data[ph_name]
                p_new = pharm_new_pat_month.get(ph_name, {})
                p_months_sorted = sorted(p_m.keys())

                # 药房名清洗：去除 Excel 工作表名非法字符（\ / * ? : [ ]），否则 create_sheet 会抛错
                raw_name = ph_name[:15] if len(ph_name) > 15 else ph_name
                sheet_name = _safe_sheet_name(raw_name)

                # 去重：如果Sheet名已存在，用清洗后的全名试，仍冲突则加序号
                if sheet_name in used_sheet_names:
                    sheet_name = _safe_sheet_name(ph_name)[:30]
                if sheet_name in used_sheet_names:
                    cnt = 1
                    base = sheet_name[:28]
                    while _safe_sheet_name(f"{base}_{cnt}") in used_sheet_names:
                        cnt += 1
                    sheet_name = _safe_sheet_name(f"{base}_{cnt}")[:31]
                used_sheet_names.add(sheet_name)

                # 新建空Sheet，手工填入所有内容
                wsP = writer.book.create_sheet(sheet_name)
                wsP_time = f"药房：{ph_name}  |  {time_label}"
                wr = 1  # 当前写入行

                # ---- 药房标题 ----
                wsP.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=11)
                c = wsP.cell(row=wr, column=1, value=f'药房分析报告 — {ph_name}')
                c.font = Font(bold=True, size=14, color='1a3a5c')
                c.alignment = Alignment(horizontal='left', vertical='center')
                wr += 1

                wsP.merge_cells(start_row=wr, start_column=1, end_row=wr, end_column=11)
                c = wsP.cell(row=wr, column=1, value=wsP_time)
                c.font = time_font; c.fill = time_fill
                c.alignment = Alignment(horizontal='left', vertical='center')
                wr += 2

                # ===== 维度1：月度趋势数据 =====
                wr = write_section_title(wsP, wr, 1, '📈 一、月度趋势数据', 6); wr += 1
                trend_rows = []
                for mk in p_months_sorted:
                    act = len(p_m[mk]['patients'])
                    new = len(p_new.get(mk, set()))
                    trend_rows.append({
                        '药房': ph_name, '月份': get_month_label(mk),
                        '购药总支数': round(p_m[mk]['qty'], 2),
                        '活跃患者数': act, '新患患者数': new,
                        '筛选期内': '是' if sm_store <= mk <= em_store else '否',
                    })
                if trend_rows:
                    tdf = pd.DataFrame(trend_rows)
                    for ci, cn in enumerate(tdf.columns, 1):
                        c = wsP.cell(row=wr, column=ci, value=cn)
                        c.font = header_font; c.fill = header_fill
                    wr += 1
                    for _, row in tdf.iterrows():
                        for ci, cn in enumerate(tdf.columns, 1):
                            wsP.cell(row=wr, column=ci, value=row[cn])
                        wr += 1
                else:
                    wsP.cell(row=wr, column=1, value='（无月度趋势数据）'); wr += 1
                wr += 2

                # ===== 维度2：月度DOT对比数据 =====
                wr = write_section_title(wsP, wr, 1, '📊 二、月度DOT对比数据', 6); wr += 1
                dot_rows = calc_rolling_dot_for_pharm(p_m)
                if dot_rows:
                    ddf = pd.DataFrame(dot_rows)
                    for ci, cn in enumerate(ddf.columns, 1):
                        c = wsP.cell(row=wr, column=ci, value=cn)
                        c.font = header_font; c.fill = header_fill
                    wr += 1
                    for _, row in ddf.iterrows():
                        for ci, cn in enumerate(ddf.columns, 1):
                            wsP.cell(row=wr, column=ci, value=row[cn])
                        wr += 1
                else:
                    wsP.cell(row=wr, column=1, value='（无DOT对比数据）'); wr += 1
                wr += 2

                # ===== 维度3：活跃vs新患对比 =====
                wr = write_section_title(wsP, wr, 1, '👥 三、活跃患者 vs 新患患者对比', 5); wr += 1
                cmp_rows = []
                for mk in p_months_sorted:
                    if sm_store <= mk <= em_store:
                        act = len(p_m[mk]['patients'])
                        new = len(p_new.get(mk, set()))
                        cmp_rows.append({
                            '药房': ph_name, '月份': get_month_label(mk),
                            '活跃患者数': act, '新患患者数': new,
                            '新患占比(%)': round(new / act * 100, 2) if act > 0 else 0,
                        })
                if cmp_rows:
                    cdf = pd.DataFrame(cmp_rows)
                    for ci, cn in enumerate(cdf.columns, 1):
                        c = wsP.cell(row=wr, column=ci, value=cn)
                        c.font = header_font; c.fill = header_fill
                    wr += 1
                    for _, row in cdf.iterrows():
                        for ci, cn in enumerate(cdf.columns, 1):
                            wsP.cell(row=wr, column=ci, value=row[cn])
                        wr += 1
                else:
                    wsP.cell(row=wr, column=1, value='（无活跃vs新患对比数据）'); wr += 1
                wr += 2

                # ===== 维度4：患者购药支数分布 =====
                wr = write_section_title(wsP, wr, 1, '📐 四、患者购药支数分布', 5); wr += 1
                pharm_patient_qty = {}
                for pid, pinfo in patient_data_local.items():
                    if ph_name in pinfo.get('pharmacy_qty', {}):
                        pharm_patient_qty[pid] = pinfo['pharmacy_qty'][ph_name]
                total_pharm_patients = len(pharm_patient_qty)
                dist_rows = []
                for label, lo, hi in buckets_def:
                    cnt = sum(1 for q in pharm_patient_qty.values() if lo <= q <= hi)
                    dist_rows.append({
                        '药房': ph_name, '购药支数区间': label,
                        '患者人数': cnt,
                        '占比(%)': round(cnt / total_pharm_patients * 100, 2) if total_pharm_patients > 0 else 0,
                        '统计时间范围': f"{get_month_label(p_months_sorted[0])}~{get_month_label(p_months_sorted[-1])}" if p_months_sorted else '-',
                    })
                if dist_rows:
                    bdf = pd.DataFrame(dist_rows)
                    for ci, cn in enumerate(bdf.columns, 1):
                        c = wsP.cell(row=wr, column=ci, value=cn)
                        c.font = header_font; c.fill = header_fill
                    wr += 1
                    for _, row in bdf.iterrows():
                        for ci, cn in enumerate(bdf.columns, 1):
                            wsP.cell(row=wr, column=ci, value=row[cn])
                        wr += 1
                else:
                    wsP.cell(row=wr, column=1, value='（无患者分布数据）'); wr += 1
                wr += 2

                # ===== 维度5：年度DOT对比 =====
                wr = write_section_title(wsP, wr, 1, '📅 五、年度DOT对比', 11); wr += 1
                annual_rows = calc_annual_dot_for_pharm(p_m, sm_store, em_store)
                for r in annual_rows:
                    r['药房'] = ph_name
                if annual_rows:
                    adf = pd.DataFrame(annual_rows)
                    for ci, cn in enumerate(adf.columns, 1):
                        c = wsP.cell(row=wr, column=ci, value=cn)
                        c.font = header_font; c.fill = header_fill
                    wr += 1
                    for _, row in adf.iterrows():
                        for ci, cn in enumerate(adf.columns, 1):
                            wsP.cell(row=wr, column=ci, value=row[cn])
                        wr += 1
                else:
                    wsP.cell(row=wr, column=1, value='（无年度DOT对比数据）'); wr += 1

                # 设置药房Sheet列宽
                ph_widths = [18, 12, 14, 14, 14, 12, 18, 18, 12, 14, 15]
                for ci, w in enumerate(ph_widths, 1):
                    wsP.column_dimensions[chr(64 + ci)].width = w

            pharm_count = len(pharm_monthly_data)

            # ---- 透视表独立下载（仅包含透视汇总数据） ----
            excel_buf_pivot = BytesIO()
            with pd.ExcelWriter(excel_buf_pivot, engine='openpyxl') as writer_pivot:
                # 【防御】透视工作簿同样保证至少一个【可见】Sheet，避免 IndexError 掩盖真实错误
                _pwbs = getattr(writer_pivot.book, 'worksheets', [])
                if not any(getattr(ws, 'sheet_state', 'visible') == 'visible' for ws in _pwbs):
                    if _pwbs:
                        _pwbs[0].sheet_state = 'visible'
                    else:
                        writer_pivot.book.create_sheet('Sheet1')
                # Sheet 1: 品种×药房透视汇总
                csdf_local = ad_local.get('cross_summary_df', pd.DataFrame())
                if not csdf_local.empty:
                    csdf_local.to_excel(writer_pivot, sheet_name='品种×药房透视汇总', index=False)
                    ws_p1 = writer_pivot.sheets['品种×药房透视汇总']
                    add_time_header(ws_p1, len(csdf_local.columns))
                    for ci, w in enumerate([8, 22, 25, 14, 14, 18, 10], 1):
                        ws_p1.column_dimensions[chr(64+ci)].width = w
                else:
                    pd.DataFrame({'提示': ['需要同时配置品种和药房字段']}).to_excel(
                        writer_pivot, sheet_name='品种×药房透视汇总', index=False)

                # Sheet 2: 药房汇总
                phsdf_local = ad_local.get('pharmacy_summary_df', pd.DataFrame())
                if not phsdf_local.empty:
                    phsdf_local.to_excel(writer_pivot, sheet_name='药房DOT汇总', index=False)
                    ws_p2 = writer_pivot.sheets['药房DOT汇总']
                    add_time_header(ws_p2, len(phsdf_local.columns))
                    for ci, w in enumerate([8, 25, 14, 14, 18, 10], 1):
                        ws_p2.column_dimensions[chr(64+ci)].width = w
                else:
                    pd.DataFrame({'提示': ['无药房汇总数据']}).to_excel(
                        writer_pivot, sheet_name='药房DOT汇总', index=False)

                # Sheet 3: 品种汇总
                psdf_local = ad_local.get('product_summary_df', pd.DataFrame())
                if not psdf_local.empty:
                    psdf_local.to_excel(writer_pivot, sheet_name='品种汇总', index=False)
                    ws_p3 = writer_pivot.sheets['品种汇总']
                    add_time_header(ws_p3, len(psdf_local.columns))
                    for ci, w in enumerate([8, 25, 14, 14, 18, 10], 1):
                        ws_p3.column_dimensions[chr(64+ci)].width = w
                else:
                    pd.DataFrame({'提示': ['未检测到品种字段或无数据']}).to_excel(
                        writer_pivot, sheet_name='品种汇总', index=False)

        st.download_button(
            label=f"📥 导出完整报告（汇总 + 全局汇总 + {pharm_count} 个药房分析）",
            data=excel_buf.getvalue(),
            file_name=f"阿斯利康数据分析_{_fmt_date(start_dt)}_{_fmt_date(end_dt)}.xlsx".replace(' ', ''),
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )

        st.download_button(
            label=f"📊 单独下载透视汇总表（品种×药房 + 药房DOT + 品种）",
            data=excel_buf_pivot.getvalue(),
            file_name=f"透视汇总_{_fmt_date(start_dt)}_{_fmt_date(end_dt)}.xlsx".replace(' ', ''),
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )

        # 兜底保障：保存前至少保留一个可见工作表，
        # 避免 openpyxl 在 save 时抛 IndexError("At least one sheet must be visible")
        # （该错误会掩盖真实的建表异常，导致难以定位）
        if not any(getattr(ws, 'sheet_state', 'visible') == 'visible' for ws in writer.book.worksheets):
            if writer.book.worksheets:
                writer.book.worksheets[0].sheet_state = 'visible'
            else:
                writer.book.create_sheet('Sheet1')

    # ==================== 数据预览 ====================
    if st.session_state.dot_result is None:
        st.markdown("---")
        st.header("👀 数据预览")
        st.dataframe(df.head(20), use_container_width=True, hide_index=True)
        st.caption(f"显示前 20 条，共 {len(df):,} 条记录")
